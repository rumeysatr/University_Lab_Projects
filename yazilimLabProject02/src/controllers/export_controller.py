"""
Dışa aktarma controller - Excel export işlemleri
View ile Service arasında köprü görevi görür
"""

from typing import List, Dict, Any, Optional
from datetime import date, datetime
import os
from src.services.exam_schedule_service import ExamScheduleService
from src.services.course_service import CourseService
from src.services.classroom_service import ClassroomService
from src.utils.excel_generator import ExcelGenerator


def validate_file_path(path: str, extension: str) -> bool:
    if not path:
        return False
    return path.lower().endswith(f'.{extension}')


class ExportController:
    
    def __init__(self, view: Any = None):
        self.view = view
        self.exam_service = ExamScheduleService()
        self.course_service = CourseService()
        self.classroom_service = ClassroomService()
    
    def export_exam_schedule_to_excel(self, start_date: date, end_date: date,
                                      output_path: str = None) -> tuple:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            if output_path and not validate_file_path(output_path, 'xlsx'):
                return False, "Geçersiz dosya yolu veya uzantısı", None
            
            exams = self.exam_service.get_by_date_range(start_date, end_date)
            if not exams:
                return False, "Belirtilen tarih aralığında sınav bulunamadı", None
            
            if not output_path:
                output_path = f"sinav_programi_{start_date}_{end_date}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sınav Programı"
            
            headers = ['Tarih', 'Saat', 'Ders Kodu', 'Ders Adı', 'Öğretim Görevlisi',
                      'Derslik', 'Öğrenci Sayısı', 'Sınav Türü', 'Durum']
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            row = 2
            for exam in exams:
                ws.cell(row=row, column=1, value=exam.exam_date.strftime('%d.%m.%Y'))
                ws.cell(row=row, column=2, value=f"{exam.start_time} - {exam.end_time}")
                ws.cell(row=row, column=3, value=exam.course_code)
                ws.cell(row=row, column=4, value=exam.course_name)
                ws.cell(row=row, column=5, value=exam.lecturer_name)
                ws.cell(row=row, column=6, value=f"{exam.faculty_name or 'Belirsiz'} - {exam.classroom_name}")
                ws.cell(row=row, column=7, value=exam.student_count)
                ws.cell(row=row, column=8, value=self._get_exam_type_label(exam.exam_type))
                ws.cell(row=row, column=9, value=self._get_status_label(exam.status))
                row += 1
            
            column_widths = [15, 15, 15, 30, 25, 20, 12, 15, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            wb.save(output_path)
            
            return True, "Excel dosyası başarıyla oluşturuldu", output_path
            
        except ImportError:
            return False, "Excel export için openpyxl kütüphanesi gereklidir", None
        except Exception as e:
            return False, f"Excel oluşturulurken hata: {str(e)}", None
    
    def _get_exam_type_label(self, exam_type: str) -> str:
        labels = {
            'midterm': 'Vize',
            'final': 'Final',
            'makeup': 'Bütünleme',
            'quiz': 'Quiz'
        }
        return labels.get(exam_type, exam_type)
    
    def _get_status_label(self, status: str) -> str:
        labels = {
            'planned': 'Planlandı',
            'confirmed': 'Onaylandı',
            'cancelled': 'İptal Edildi'
        }
        return labels.get(status, status)
    
    def get_export_templates(self) -> Dict[str, str]:
        return {
            'exam_schedule_excel': 'Sınav Programı (Excel)',
        }
    
    def export_to_excel(self, data: list = None, file_path: str = None,
                        filter_type: str = None, filter_value: int = None) -> dict:
        try:
            if data is None:
                if filter_type == 'faculty':
                    data = self.exam_service.get_by_faculty_id(filter_value)
                elif filter_type == 'department':
                    data = self.exam_service.get_by_department_id(filter_value)
                elif filter_type == 'classroom':
                    data = self.exam_service.get_by_classroom_id(filter_value)
                else:
                    data = self.exam_service.get_all()
            
            if not data:
                return {
                    'success': False,
                    'message': 'Dışa aktarılacak veri bulunamadı',
                    'path': None
                }
            
            if not file_path:
                file_path = f"rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            generator = ExcelGenerator()
            success = generator.generate(data, file_path)
            
            return {
                'success': success,
                'message': 'Excel oluşturuldu' if success else 'Excel oluşturulamadı',
                'path': file_path if success else None
            }
            
        except Exception as e:
            return {'success': False, 'message': str(e), 'path': None}
