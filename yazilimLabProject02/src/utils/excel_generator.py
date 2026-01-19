"""
Excel Generator - Excel dosyası oluşturma araçları
Dict ve Object destekli
"""

from typing import List, Any, Optional
from datetime import date


class ExcelGenerator:    
    def __init__(self):
        self.title = "Sınav Programı"
        self.author = "Sınav Planlama Sistemi"
    
    def _get_value(self, item: Any, key: str, default: Any = '') -> Any:
        if isinstance(item, dict):
            return item.get(key, default)
        else:
            return getattr(item, key, default)
    
    def generate(self, data: List[Any], output_path: str) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = self.title[:31]
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if not data:
                wb.save(output_path)
                return True
            
            first_item = data[0]

            if self._get_value(first_item, 'type') == 'classroom_conflict' or self._get_value(first_item, 'course1'):
                return self._generate_conflict_excel(data, output_path, wb, ws, header_fill, header_font, thin_border)
            elif self._get_value(first_item, 'istatistik'):
                return self._generate_statistics_excel(data, output_path, wb, ws, header_fill, header_font, thin_border)
            elif self._get_value(first_item, 'zaman_dilimi'):
                return self._generate_classroom_usage_excel(data, output_path, wb, ws, header_fill, header_font, thin_border)
            elif self._get_value(first_item, 'faculty_name') and self._get_value(first_item, 'department_count'):
                return self._generate_faculty_excel(data, output_path, wb, ws, header_fill, header_font, thin_border)
            elif self._get_value(first_item, 'course_code'):
                return self._generate_exam_dict_excel(data, output_path, wb, ws, header_fill, header_font, thin_border)
            else:
                return self._generate_generic_excel(data, output_path, wb, ws, header_fill, header_font, thin_border)
            
        except ImportError:
            return self._fallback_to_csv(data, output_path)
        except Exception as e:
            print(f"Excel oluşturma hatası: {e}")
            return self._fallback_to_csv(data, output_path)
    
    def _generate_exam_excel(self, data, output_path, wb, ws, header_fill, header_font, thin_border) -> bool:
        from openpyxl.styles import Alignment
        
        headers = ['Tarih', 'Başlangıç Saati', 'Bitiş Saati', 'Ders Kodu',
                  'Ders Adı', 'Öğretim Üyesi', 'Derslik', 'Fakülte',
                  'Öğrenci Sayısı', 'Sınav Türü', 'Durum']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2
        for item in data:
            ws.cell(row=row, column=1, value=str(self._get_value(item, 'exam_date', '')))
            ws.cell(row=row, column=2, value=str(self._get_value(item, 'start_time', '')))
            ws.cell(row=row, column=3, value=str(self._get_value(item, 'end_time', '')))
            ws.cell(row=row, column=4, value=str(self._get_value(item, 'course_code', '')))
            ws.cell(row=row, column=5, value=str(self._get_value(item, 'course_name', '')))
            ws.cell(row=row, column=6, value=str(self._get_value(item, 'lecturer_name', '')))
            ws.cell(row=row, column=7, value=str(self._get_value(item, 'classroom_name', '')))
            ws.cell(row=row, column=8, value=str(self._get_value(item, 'faculty_name', '') or 'Belirsiz'))
            ws.cell(row=row, column=9, value=str(self._get_value(item, 'student_count', '')))
            ws.cell(row=row, column=10, value=self._get_exam_type_label(self._get_value(item, 'exam_type', '')))
            ws.cell(row=row, column=11, value=self._get_status_label(self._get_value(item, 'status', '')))
            
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = thin_border
            
            row += 1
        
        column_widths = [12, 12, 12, 12, 25, 20, 15, 15, 12, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[self._get_column_letter(i)].width = width
        
        wb.save(output_path)
        return True
    
    def _generate_exam_dict_excel(self, data, output_path, wb, ws, header_fill, header_font, thin_border) -> bool:
        from openpyxl.styles import Alignment
        
        headers = ['Tarih', 'Başlangıç Saati', 'Bitiş Saati', 'Ders Kodu', 'Ders Adı',
                   'Öğretim Üyesi', 'Derslik', 'Fakülte', 'Bölüm', 'Öğrenci Sayısı', 'Sınav Türü', 'Durum']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2
        for item in data:
            exam_date = self._get_value(item, 'date', '') or self._get_value(item, 'exam_date', '')
            ws.cell(row=row, column=1, value=str(exam_date))
            
            start_time = self._get_value(item, 'start_time', '') or self._get_value(item, 'time', '')
            ws.cell(row=row, column=2, value=str(start_time))
            
            end_time = self._get_value(item, 'end_time', '')
            ws.cell(row=row, column=3, value=str(end_time))
            
            ws.cell(row=row, column=4, value=str(self._get_value(item, 'course_code', '')))
            ws.cell(row=row, column=5, value=str(self._get_value(item, 'course_name', '')))
            
            lecturer = self._get_value(item, 'lecturer_name', '') or self._get_value(item, 'lecturer', '')
            ws.cell(row=row, column=6, value=str(lecturer))
            
            classroom = self._get_value(item, 'classroom_name', '') or self._get_value(item, 'classroom', '')
            ws.cell(row=row, column=7, value=str(classroom))
            
            faculty = self._get_value(item, 'faculty_name', '')
            ws.cell(row=row, column=8, value=str(faculty))
            
            department = self._get_value(item, 'department_name', '')
            ws.cell(row=row, column=9, value=str(department))
            
            ws.cell(row=row, column=10, value=str(self._get_value(item, 'student_count', '')))
            
            exam_type = self._get_value(item, 'exam_type', '')
            ws.cell(row=row, column=11, value=self._get_exam_type_label(exam_type))
            
            status = self._get_value(item, 'status', '')
            ws.cell(row=row, column=12, value=self._get_status_label(status))
            
            for col in range(1, 13):
                ws.cell(row=row, column=col).border = thin_border
            
            row += 1
        
        column_widths = [12, 12, 12, 12, 25, 20, 15, 15, 18, 12, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[self._get_column_letter(i)].width = width
        
        wb.save(output_path)
        return True
    
    def _generate_faculty_excel(self, data, output_path, wb, ws, header_fill, header_font, thin_border) -> bool:
        from openpyxl.styles import Alignment
        
        headers = ['Fakülte Adı', 'Fakülte Kodu', 'Bölüm Sayısı']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2
        for item in data:
            ws.cell(row=row, column=1, value=str(self._get_value(item, 'faculty_name', '')))
            ws.cell(row=row, column=2, value=str(self._get_value(item, 'code', '')))
            ws.cell(row=row, column=3, value=str(self._get_value(item, 'department_count', '')))
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = thin_border
            
            row += 1
        
        for i, width in enumerate([30, 15, 15], 1):
            ws.column_dimensions[self._get_column_letter(i)].width = width
        
        wb.save(output_path)
        return True
    
    def _generate_classroom_usage_excel(self, data, output_path, wb, ws, header_fill, header_font, thin_border) -> bool:
        from openpyxl.styles import Alignment
        
        headers = ['Zaman Dilimi', 'Kullanılan Derslik', 'Toplam Derslik', 'Kullanılan Kapasite', 'Toplam Kapasite', 'Kullanım %']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2
        for item in data:
            ws.cell(row=row, column=1, value=str(self._get_value(item, 'zaman_dilimi', '')))
            ws.cell(row=row, column=2, value=str(self._get_value(item, 'used_classrooms', '')))
            ws.cell(row=row, column=3, value=str(self._get_value(item, 'total_classrooms', '')))
            ws.cell(row=row, column=4, value=str(self._get_value(item, 'used_capacity', '')))
            ws.cell(row=row, column=5, value=str(self._get_value(item, 'total_capacity', '')))
            percentage = self._get_value(item, 'percentage', 0)
            ws.cell(row=row, column=6, value=f"{percentage:.1f}%" if isinstance(percentage, (int, float)) else str(percentage))
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
            
            row += 1
        
        for i, width in enumerate([15, 18, 15, 18, 15, 12], 1):
            ws.column_dimensions[self._get_column_letter(i)].width = width
        
        wb.save(output_path)
        return True
    
    def _generate_statistics_excel(self, data, output_path, wb, ws, header_fill, header_font, thin_border) -> bool:
        from openpyxl.styles import Alignment
        
        stat_labels = {
            'total_faculties': 'Toplam Fakülte',
            'total_departments': 'Toplam Bölüm',
            'total_classrooms': 'Toplam Derslik',
            'total_lecturers': 'Toplam Öğretim Üyesi',
            'total_courses': 'Toplam Ders',
            'total_exams': 'Toplam Sınav',
            'this_week_exams': 'Bu Haftaki Sınavlar',
            'today_exams': 'Bugünkü Sınavlar',
            'pending_exams': 'Onay Bekleyen',
        }
        
        headers = ['İstatistik', 'Değer']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2
        for item in data:
            stat_key = self._get_value(item, 'istatistik', '')
            stat_value = self._get_value(item, 'deger', '')
            stat_label = stat_labels.get(stat_key, stat_key)
            
            ws.cell(row=row, column=1, value=stat_label)
            ws.cell(row=row, column=2, value=str(stat_value))
            
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = thin_border
            
            row += 1
        
        for i, width in enumerate([25, 15], 1):
            ws.column_dimensions[self._get_column_letter(i)].width = width
        
        wb.save(output_path)
        return True
    
    def _generate_conflict_excel(self, data, output_path, wb, ws, header_fill, header_font, thin_border) -> bool:
        from openpyxl.styles import Alignment
        
        headers = ['Tarih', 'Saat', 'Derslik', 'Ders 1', 'Ders 2', 'Çakışma Tipi']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2
        for item in data:
            ws.cell(row=row, column=1, value=str(self._get_value(item, 'date', '')))
            ws.cell(row=row, column=2, value=str(self._get_value(item, 'time', '')))
            ws.cell(row=row, column=3, value=str(self._get_value(item, 'classroom', '')))
            ws.cell(row=row, column=4, value=str(self._get_value(item, 'course1', '')))
            ws.cell(row=row, column=5, value=str(self._get_value(item, 'course2', '')))
            conflict_type = self._get_value(item, 'type', 'classroom_conflict')
            ws.cell(row=row, column=6, value='Derslik Çakışması' if conflict_type == 'classroom_conflict' else str(conflict_type))
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
            
            row += 1
        
        for i, width in enumerate([12, 15, 15, 15, 15, 18], 1):
            ws.column_dimensions[self._get_column_letter(i)].width = width
        
        wb.save(output_path)
        return True
    
    def _generate_generic_excel(self, data, output_path, wb, ws, header_fill, header_font, thin_border) -> bool:
        from openpyxl.styles import Alignment
        
        if not data:
            wb.save(output_path)
            return True
        
        first_item = data[0]
        
        if isinstance(first_item, dict):
            keys = list(first_item.keys())
        else:
            keys = [attr for attr in dir(first_item) 
                    if not attr.startswith('_') and not callable(getattr(first_item, attr))]
        
        for col, key in enumerate(keys, 1):
            cell = ws.cell(row=1, column=col, value=str(key))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row = 2
        for item in data:
            for col, key in enumerate(keys, 1):
                ws.cell(row=row, column=col, value=str(self._get_value(item, key, '')))
                ws.cell(row=row, column=col).border = thin_border
            row += 1
        
        for i in range(1, len(keys) + 1):
            ws.column_dimensions[self._get_column_letter(i)].width = 15
        
        wb.save(output_path)
        return True
    
    def _get_column_letter(self, col_num: int) -> str:
        result = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            result = chr(65 + remainder) + result
        return result
    
    def _get_exam_type_label(self, exam_type: str) -> str:
        labels = {
            'midterm': 'Vize',
            'final': 'Final',
            'makeup': 'Bütünleme',
            'quiz': 'Quiz'
        }
        return labels.get(exam_type, exam_type or '')
    
    def _get_status_label(self, status: str) -> str:
        labels = {
            'planned': 'Planlandı',
            'confirmed': 'Onaylandı',
            'cancelled': 'İptal Edildi'
        }
        return labels.get(status, status or '')
    
    def _fallback_to_csv(self, data: List[Any], output_path: str) -> bool:
        try:
            import csv
            csv_path = output_path.replace('.xlsx', '.csv')
            
            if not data:
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    f.write('')
                return True
            
            first_item = data[0]
            
            if isinstance(first_item, dict):
                keys = list(first_item.keys())
            else:
                keys = [attr for attr in dir(first_item) 
                        if not attr.startswith('_') and not callable(getattr(first_item, attr))]
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                writer.writerow(keys)
                
                for item in data:
                    row = [str(self._get_value(item, key, '')) for key in keys]
                    writer.writerow(row)
            
            print(f"Not: Excel oluşturulamadı, CSV olarak kaydedildi: {csv_path}")
            return True
        except Exception as e:
            print(f"CSV oluşturma hatası: {e}")
            return False
    
    def generate_classroom_list(self, classrooms: List[Any], output_path: str) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Derslikler"
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            headers = ['ID', 'Derslik Adı', 'Fakülte', 'Kapasite', 'Bilgisayar']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            row = 2
            for classroom in classrooms:
                ws.cell(row=row, column=1, value=self._get_value(classroom, 'id', ''))
                ws.cell(row=row, column=2, value=str(self._get_value(classroom, 'name', '')))
                ws.cell(row=row, column=3, value=str(self._get_value(classroom, 'faculty_name', '') or 'Belirsiz'))
                ws.cell(row=row, column=4, value=self._get_value(classroom, 'capacity', 0))
                has_computer = self._get_value(classroom, 'has_computer', False)
                ws.cell(row=row, column=5, value='Var' if has_computer else 'Yok')
                row += 1
            
            for i, width in enumerate([8, 20, 20, 12, 10], 1):
                ws.column_dimensions[self._get_column_letter(i)].width = width
            
            wb.save(output_path)
            return True
            
        except ImportError:
            return False
        except Exception as e:
            print(f"Derslik listesi Excel hatası: {e}")
            return False
    
    def generate_course_list(self, courses: List[Any], output_path: str) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Dersler"
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            headers = ['Ders Kodu', 'Ders Adı', 'Kredi', 'AKTS', 'Dönem', 
                      'Öğrenci Sayısı', 'Öğretim Üyesi', 'Bölüm']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            row = 2
            for course in courses:
                ws.cell(row=row, column=1, value=str(self._get_value(course, 'code', '')))
                ws.cell(row=row, column=2, value=str(self._get_value(course, 'name', '')))
                ws.cell(row=row, column=3, value=self._get_value(course, 'credit', 0))
                ws.cell(row=row, column=4, value=self._get_value(course, 'ects', 0))
                ws.cell(row=row, column=5, value=self._get_value(course, 'semester', ''))
                ws.cell(row=row, column=6, value=self._get_value(course, 'student_count', 0))
                ws.cell(row=row, column=7, value=str(self._get_value(course, 'lecturer_name', '')))
                ws.cell(row=row, column=8, value=str(self._get_value(course, 'department_name', '')))
                row += 1
            
            for i, width in enumerate([12, 30, 8, 8, 8, 12, 20, 20], 1):
                ws.column_dimensions[self._get_column_letter(i)].width = width
            
            wb.save(output_path)
            return True
            
        except ImportError:
            return False
        except Exception as e:
            print(f"Ders listesi Excel hatası: {e}")
            return False
