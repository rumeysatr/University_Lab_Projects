"""
Excel kapasite dosyası oluşturucu script.

Bu script kostu_sinav_kapasiteleri.xlsx dosyasını oluşturur.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


def create_capacity_excel():
    """Sınıf kapasite Excel dosyasını oluşturur."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kapasite"

    # Başlıkları ekle
    headers = ["Blok", "Derslik Numarası", "Kapasite (Kişi)"]
    ws.append(headers)

    # Başlık stilini uygula
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Sınıf verilerini ekle (20+ derslik)
    classrooms = [
        ("M", "M101", 50),
        ("M", "M102", 60),
        ("M", "M103", 45),
        ("M", "M104", 55),
        ("M", "M201", 40),
        ("S", "S101", 80),
        ("S", "S102", 75),
        ("S", "S103", 90),
        ("S", "S104", 85),
        ("S", "S201", 70),
        ("E", "E101", 35),
        ("E", "E102", 40),
        ("E", "E103", 38),
        ("E", "E104", 42),
        ("E", "E201", 36),
        ("K", "K101", 50),
        ("K", "K102", 48),
        ("K", "K103", 52),
        ("K", "K104", 46),
        ("K", "K201", 50),
        ("N", "N101", 65),
        ("N", "N102", 70),
        ("N", "N103", 68),
    ]

    # Veriyi ekle
    center_alignment = Alignment(horizontal="center", vertical="center")
    for blok, derslik, kapasite in classrooms:
        row = [blok, derslik, kapasite]
        ws.append(row)
        
        # Son eklenen satıra stili uygula
        for cell in ws[ws.max_row]:
            cell.alignment = center_alignment
            cell.border = border

    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    # Dosyayı kaydet
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(script_dir))
    excel_dir = os.path.join(project_root, 'database', 'exceller')
    os.makedirs(excel_dir, exist_ok=True)
    
    output_path = os.path.join(excel_dir, 'kostu_sinav_kapasiteleri.xlsx')
    wb.save(output_path)
    print(f"✅ kostu_sinav_kapasiteleri.xlsx başarıyla oluşturuldu: {output_path}")
    return output_path


def create_proximity_excel():
    """Derslik yakınlık tablosu Excel dosyasını oluşturur."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yakinlik"

    # Başlıkları ekle
    headers = ["Derslik 1", "Derslik 2", "Mesafe (km)", "Yakınlık Derecesi"]
    ws.append(headers)

    # Başlık stilini uygula
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Derslik yakınlıkları (Yakınlık Derecesi: 1=çok yakın, 5=çok uzak)
    proximities = [
        ("M101", "M102", 0.01, 1),
        ("M101", "M103", 0.02, 1),
        ("M101", "M104", 0.03, 1),
        ("M101", "M201", 0.05, 2),
        ("M101", "S101", 0.15, 3),
        ("M101", "S102", 0.16, 3),
        ("M101", "E101", 0.25, 4),
        ("M101", "K101", 0.35, 5),
        ("M102", "M103", 0.01, 1),
        ("M102", "M104", 0.02, 1),
        ("M102", "M201", 0.04, 2),
        ("M102", "S101", 0.14, 3),
        ("S101", "S102", 0.01, 1),
        ("S101", "S103", 0.02, 1),
        ("S101", "S104", 0.03, 1),
        ("S101", "S201", 0.05, 2),
        ("S101", "E101", 0.20, 4),
        ("S101", "K101", 0.30, 4),
        ("E101", "E102", 0.01, 1),
        ("E101", "E103", 0.02, 1),
        ("E101", "E104", 0.03, 1),
        ("E101", "E201", 0.05, 2),
        ("K101", "K102", 0.01, 1),
        ("K101", "K103", 0.02, 1),
        ("K101", "K104", 0.03, 1),
        ("K101", "K201", 0.05, 2),
        ("N101", "N102", 0.01, 1),
        ("N101", "N103", 0.02, 1),
        ("M101", "N101", 0.40, 5),
        ("S101", "N101", 0.35, 5),
    ]

    # Veriyi ekle
    center_alignment = Alignment(horizontal="center", vertical="center")
    for derslik1, derslik2, mesafe, yakinlik in proximities:
        row = [derslik1, derslik2, mesafe, yakinlik]
        ws.append(row)
        
        # Son eklenen satıra stili uygula
        for cell in ws[ws.max_row]:
            cell.alignment = center_alignment
            cell.border = border

    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18

    # Dosyayı kaydet
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(script_dir))
    excel_dir = os.path.join(project_root, 'database', 'exceller')
    os.makedirs(excel_dir, exist_ok=True)
    
    output_path = os.path.join(excel_dir, 'DerslikYakinlik.xlsx')
    wb.save(output_path)
    print(f"✅ DerslikYakinlik.xlsx başarıyla oluşturuldu: {output_path}")
    return output_path


if __name__ == '__main__':
    create_capacity_excel()
    create_proximity_excel()
