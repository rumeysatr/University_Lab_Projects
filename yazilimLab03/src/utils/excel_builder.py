"""
Excel dosyaları oluşturmak için utilities.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Tuple, Dict
import os


class ExcelBuilder:
    """Excel dosyaları oluşturmak için helper sınıfı."""

    # Sabit stiller
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def apply_header_style(ws, headers: List[str]):
        """Başlık stilini uygular."""
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = ExcelBuilder.HEADER_FILL
            cell.font = ExcelBuilder.HEADER_FONT
            cell.alignment = ExcelBuilder.CENTER_ALIGN
            cell.border = ExcelBuilder.BORDER

    @staticmethod
    def add_row(ws, row: List, centered: bool = True):
        """Satır ekler ve stili uygular."""
        ws.append(row)
        alignment = ExcelBuilder.CENTER_ALIGN if centered else Alignment(vertical="center")
        for cell in ws[ws.max_row]:
            cell.alignment = alignment
            cell.border = ExcelBuilder.BORDER

    @staticmethod
    def create_capacity_file(output_path: str):
        """Sınıf kapasite dosyasını oluşturur."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kapasite"

        # Başlıkları ekle
        headers = ["Blok", "Derslik Numarası", "Kapasite (Kişi)"]
        ExcelBuilder.apply_header_style(ws, headers)

        # Derslik verilerini ekle
        classrooms = [
            ("M", "M101", 50),
            ("M", "M102", 60),
            ("M", "M103", 45),
            ("M", "M104", 55),
            ("M", "M201", 40),
            ("S", "S101", 80),
            ("S", "S102", 75),
            ("S", "S103", 90),
            ("S", "S104", 85),
            ("S", "S201", 70),
            ("E", "E101", 35),
            ("E", "E102", 40),
            ("E", "E103", 38),
            ("E", "E104", 42),
            ("E", "E201", 36),
            ("K", "K101", 50),
            ("K", "K102", 48),
            ("K", "K103", 52),
            ("K", "K104", 46),
            ("K", "K201", 50),
            ("N", "N101", 65),
            ("N", "N102", 70),
            ("N", "N103", 68),
        ]

        for blok, derslik, kapasite in classrooms:
            ExcelBuilder.add_row(ws, [blok, derslik, kapasite])

        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return output_path

    @staticmethod
    def create_proximity_file(output_path: str):
        """Derslik yakınlık dosyasını oluşturur."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Yakinlik"

        # Başlıkları ekle
        headers = ["Derslik 1", "Derslik 2", "Mesafe (km)", "Yakınlık Derecesi"]
        ExcelBuilder.apply_header_style(ws, headers)

        # Yakınlık verilerini ekle
        proximities = [
            ("M101", "M102", 0.01, 1),
            ("M101", "M103", 0.02, 1),
            ("M101", "M104", 0.03, 1),
            ("M101", "M201", 0.05, 2),
            ("M101", "S101", 0.15, 3),
            ("M101", "S102", 0.16, 3),
            ("M101", "E101", 0.25, 4),
            ("M101", "K101", 0.35, 5),
            ("M102", "M103", 0.01, 1),
            ("M102", "M104", 0.02, 1),
            ("M102", "M201", 0.04, 2),
            ("M102", "S101", 0.14, 3),
            ("S101", "S102", 0.01, 1),
            ("S101", "S103", 0.02, 1),
            ("S101", "S104", 0.03, 1),
            ("S101", "S201", 0.05, 2),
            ("S101", "E101", 0.20, 4),
            ("S101", "K101", 0.30, 4),
            ("E101", "E102", 0.01, 1),
            ("E101", "E103", 0.02, 1),
            ("E101", "E104", 0.03, 1),
            ("E101", "E201", 0.05, 2),
            ("K101", "K102", 0.01, 1),
            ("K101", "K103", 0.02, 1),
            ("K101", "K104", 0.03, 1),
            ("K101", "K201", 0.05, 2),
            ("N101", "N102", 0.01, 1),
            ("N101", "N103", 0.02, 1),
            ("M101", "N101", 0.40, 5),
            ("S101", "N101", 0.35, 5),
        ]

        for derslik1, derslik2, mesafe, yakinlik in proximities:
            ExcelBuilder.add_row(ws, [derslik1, derslik2, mesafe, yakinlik])

        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return output_path
