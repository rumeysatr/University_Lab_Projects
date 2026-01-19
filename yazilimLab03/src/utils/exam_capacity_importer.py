"""
Sınav Kapasitesi Importer
kostu_sinav_kapasiteleri.xlsx dosyasından derslik kapasite bilgilerini okur
ve ClassroomRepository üzerinden günceller.
"""

import os
import logging
from typing import Dict, List, Optional
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class CapacityData:
    """Derslik kapasite verisi"""
    block: str
    classroom_name: str
    capacity: int
    
    def __post_init__(self):
        if self.capacity <= 0:
            raise ValueError(f"Kapasite pozitif olmalı: {self.classroom_name}")


@dataclass
class CapacityImportResult:
    """Kapasite import sonucu"""
    success: bool
    message: str
    total_classrooms: int = 0
    updated_classrooms: int = 0
    new_classrooms: int = 0
    failed_count: int = 0
    errors: List[str] = None
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []


class ExamCapacityImporter:
    """Sınav kapasitesi Excel dosyası import servisi"""
    
    DEFAULT_FILE_PATH = os.path.join("database", "exceller", "kostu_sinav_kapasiteleri.xlsx")
    
    def __init__(self, classroom_repository=None):
        """
        Args:
            classroom_repository: ClassroomRepository instance (opsiyonel)
        """
        self._classroom_repo = classroom_repository
        self._repository = None
    
    @property
    def classroom_repo(self):
        """Lazy load classroom repository"""
        if self._repository is None and self._classroom_repo is not None:
            self._repository = self._classroom_repo
        elif self._repository is None:
            from src.repositories.classroom_repository import ClassroomRepository
            self._repository = ClassroomRepository()
        return self._repository
    
    def import_from_excel(self, file_path: str = None) -> CapacityImportResult:
        """
        Excel dosyasından kapasite verilerini içe aktarır.
        
        Args:
            file_path: Excel dosyası yolu. None ise varsayılan path kullanılır.
            
        Returns:
            CapacityImportResult: Import sonucu
        """
        if file_path is None:
            file_path = self._resolve_default_path()
        
        if not os.path.exists(file_path):
            return CapacityImportResult(
                success=False,
                message=f"Dosya bulunamadı: {file_path}",
                errors=[f"Dosya bulunamadı: {file_path}"]
            )
        
        try:
            capacity_data = self._load_excel_file(file_path)
        except Exception as e:
            logger.error(f"Excel dosyası okunamadı: {e}")
            return CapacityImportResult(
                success=False,
                message=f"Excel dosyası okunamadı: {str(e)}",
                errors=[str(e)]
            )
        
        if not capacity_data:
            return CapacityImportResult(
                success=False,
                message="Excel dosyasında geçerli veri bulunamadı",
                errors=["Excel dosyasında geçerli veri bulunamadı"]
            )
        
        return self._update_classroom_capacities(capacity_data)
    
    def _resolve_default_path(self) -> str:
        """Varsayılan dosya yolunu çözümle"""
        # src/utils/ dizininden projenin kök dizinine çık
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        path = os.path.join(base_dir, "database", "exceller", "kostu_sinav_kapasiteleri.xlsx")
        return os.path.normpath(path)
    
    def _load_excel_file(self, file_path: str) -> List[CapacityData]:
        """Excel dosyasını okur ve kapasite verilerini döndürür"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Başlık satırını bul
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(h).strip().lower() if h else '' for h in row]
                break
            
            # Sütun indekslerini bul
            col_map = {}
            for idx, header in enumerate(headers):
                if 'blok' in header or 'block' in header:
                    col_map['block'] = idx
                elif 'derslik' in header and 'numara' in header or 'derslik no' in header or 'classroom' in header:
                    col_map['classroom'] = idx
                elif 'kapasite' in header or 'capacity' in header or 'kişi' in header:
                    col_map['capacity'] = idx
            
            # Eğer başlık bulunamazsa varsayılan sütun sırasını kullan
            if not col_map:
                col_map = {'block': 0, 'classroom': 1, 'capacity': 2}
                logger.warning("Başlık bulunamadı, varsayılan sütun sırası kullanılıyor")
            
            capacity_data = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(c is None for c in row):
                    continue
                
                try:
                    block = str(row[col_map.get('block', 0)]).strip() if col_map.get('block') is not None else ''
                    classroom_name = str(row[col_map.get('classroom', 1)]).strip() if col_map.get('classroom') is not None else ''
                    capacity_str = row[col_map.get('capacity', 2)] if col_map.get('capacity') is not None else '0'
                    
                    if not classroom_name:
                        continue
                    
                    # Kapasiteyi integer'a çevir
                    try:
                        capacity = int(float(capacity_str)) if capacity_str else 0
                    except (ValueError, TypeError):
                        capacity = 0
                    
                    if capacity <= 0:
                        logger.warning(f"Geçersiz kapasite: {classroom_name} -> {capacity_str}")
                        continue
                    
                    capacity_data.append(CapacityData(
                        block=block,
                        classroom_name=classroom_name,
                        capacity=capacity
                    ))
                    
                except Exception as e:
                    logger.warning(f"Satır işlenirken hata: {row} -> {e}")
                    continue
            
            wb.close()
            logger.info(f"{len(capacity_data)} adet kapasite verisi okundu")
            return capacity_data
            
        except ImportError:
            # openpyxl yoksa pandas ile dene
            return self._load_with_pandas(file_path)
        except Exception as e:
            logger.error(f"Excel okuma hatası: {e}")
            raise
    
    def _load_with_pandas(self, file_path: str) -> List[CapacityData]:
        """Pandas ile Excel dosyasını okur"""
        try:
            import pandas as pd
            
            df = pd.read_excel(file_path, sheet_name=0)
            
            # Sütun adlarını normalize et
            df.columns = [str(c).strip().lower() for c in df.columns]
            
            capacity_data = []
            
            for _, row in df.iterrows():
                try:
                    # Sütunları farklı olası isimlerle kontrol et
                    block = ''
                    classroom_name = ''
                    capacity = 0
                    
                    for col in df.columns:
                        val = row.get(col)
                        if val is None or (isinstance(val, float) and pd.isna(val)):
                            continue
                        
                        val_str = str(val).strip()
                        
                        if 'blok' in col or col == 'block':
                            block = val_str
                        elif 'derslik' in col and 'numara' in col:
                            classroom_name = val_str
                        elif 'kapasite' in col or 'kişi' in col:
                            try:
                                capacity = int(float(val_str))
                            except (ValueError, TypeError):
                                pass
                    
                    if not classroom_name or capacity <= 0:
                        continue
                    
                    capacity_data.append(CapacityData(
                        block=block,
                        classroom_name=classroom_name,
                        capacity=capacity
                    ))
                    
                except Exception as e:
                    logger.warning(f"Satır işlenirken hata: {e}")
                    continue
            
            logger.info(f"Pandas ile {len(capacity_data)} adet kapasite verisi okundu")
            return capacity_data
            
        except ImportError:
            raise ImportError("Excel okumak için openpyxl veya pandas gereklidir")
        except Exception as e:
            logger.error(f"Pandas ile Excel okuma hatası: {e}")
            raise
    
    def _update_classroom_capacities(self, capacity_data: List[CapacityData]) -> CapacityImportResult:
        """Classroom kapasitelerini günceller"""
        updated_count = 0
        new_count = 0
        failed_count = 0
        errors = []
        
        existing_classrooms = self.classroom_repo.get_all()
        existing_names = {c.name: c for c in existing_classrooms}
        
        for data in capacity_data:
            try:
                if data.classroom_name in existing_names:
                    # Mevcut dersliği güncelle
                    classroom = existing_names[data.classroom_name]
                    classroom.capacity = data.capacity
                    classroom.block = data.block  # Blok bilgisini de güncelle
                    success = self.classroom_repo.update(classroom)
                    
                    if success:
                        updated_count += 1
                    else:
                        failed_count += 1
                        errors.append(f"{data.classroom_name}: Güncelleme başarısız")
                else:
                    # Yeni derslik oluştur (varsayılan faculty_id ile)
                    # Öncelikle blok bilgisi varsa o bloktaki bir fakülteyi bul
                    from src.models.classroom import Classroom
                    
                    faculty_id = self._get_faculty_id_for_block(data.block)
                    
                    # Classroom nesnesi oluştur
                    new_classroom = Classroom(
                        name=data.classroom_name,
                        faculty_id=faculty_id,
                        capacity=data.capacity,
                        has_computer=False,
                        is_suitable=True,
                        room_type=None,
                        block=data.block  # Blok bilgisini ekle
                    )
                    
                    classroom_id = self.classroom_repo.create(new_classroom)
                    success = (classroom_id > 0)
                    message = "Derslik oluşturuldu" if success else "Derslik oluşturulamadı"
                    
                    if success:
                        new_count += 1
                    else:
                        failed_count += 1
                        errors.append(f"{data.classroom_name}: {message}")
                        
            except Exception as e:
                failed_count += 1
                errors.append(f"{data.classroom_name}: {str(e)}")
                logger.error(f"Derslik güncellenirken hata: {data.classroom_name} -> {e}")
        
        total = len(capacity_data)
        success = (failed_count == 0)
        
        if success:
            message = f"Kapasite verileri başarıyla içe aktarıldı. {updated_count} derslik güncellendi, {new_count} yeni derslik oluşturuldu."
        else:
            message = f"Kapasite verileri kısmen içe aktarıldı. {updated_count} güncellendi, {new_count} yeni, {failed_count} başarısız."
        
        return CapacityImportResult(
            success=success,
            message=message,
            total_classrooms=total,
            updated_classrooms=updated_count,
            new_classrooms=new_count,
            failed_count=failed_count,
            errors=errors if errors else []
        )
    
    def _get_faculty_id_for_block(self, block: str) -> int:
        """Bloka uygun bir faculty_id döndürür. Mevcut fakülteleri kullanır veya yeni oluşturur."""
        try:
            from src.repositories.faculty_repository import FacultyRepository
            from src.models.faculty import Faculty
            faculty_repo = FacultyRepository()
            
            if not block or block.strip() == '':
                # Blok boşsa, mevcut ilk fakülteyi kullan veya oluştur
                faculties = faculty_repo.get_all()
                if faculties:
                    return faculties[0].id
                default_faculty = Faculty(
                    name="Genel Fakülte",
                    code="GENEL"
                )
                return faculty_repo.create(default_faculty)
            
            block_upper = block.upper().strip()
            
            # Önce bu blok koduna sahip fakülteyi ara
            existing_faculty = faculty_repo.get_by_code(block_upper)
            if existing_faculty:
                return existing_faculty.id
            
            # Blok için yeni bir fakülte oluştur
            new_faculty = Faculty(
                name=f"{block_upper} Bloğu Fakültesi",
                code=block_upper
            )
            new_id = faculty_repo.create(new_faculty)
            logger.info(f"Blok '{block_upper}' için yeni fakülte oluşturuldu: {new_faculty.name} (ID: {new_id})")
            return new_id
            
        except Exception as e:
            logger.warning(f"Fakülte alınırken hata oluştu, varsayılan kullanılıyor: {e}")
            # Hata durumunda mevcut fakülteleri tekrar dene
            try:
                from src.repositories.faculty_repository import FacultyRepository
                faculty_repo = FacultyRepository()
                faculties = faculty_repo.get_all()
                if faculties:
                    return faculties[0].id
            except:
                pass
            return 1


def get_capacity_importer(classroom_repository=None) -> ExamCapacityImporter:
    """Singleton kapasite importer instance döndürür"""
    return ExamCapacityImporter(classroom_repository)
