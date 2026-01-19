"""
Excel'den Öğrenci Listesi İçe Aktarma Sınıfı
"""

import os
import logging
from typing import List, Dict, Tuple, Optional, Set
from dataclasses import dataclass
from datetime import datetime
import re

import openpyxl
import xlrd

from src.models.student import Student, StudentCourse
from src.models.course import Course
from src.models.department import Department
from src.models.lecturer import Lecturer, DEFAULT_AVAILABLE_DAYS
from src.repositories.student_repository import StudentRepository, StudentCourseRepository
from src.repositories.course_repository import CourseRepository
from src.repositories.department_repository import DepartmentRepository
from src.repositories.lecturer_repository import LecturerRepository

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """İçe aktarma sonucu"""
    success: bool
    message: str
    students_imported: int = 0
    student_courses_created: int = 0
    errors: List[str] = None
    warnings: List[str] = None
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []
        if self.warnings is None:
            self.warnings = []


class StudentImporter:
    
    TURKISH_CHAR_MAP = {
        'ı': 'i', 'İ': 'I',
        'ğ': 'g', 'Ğ': 'G',
        'ü': 'u', 'Ü': 'U',
        'ş': 's', 'Ş': 'S',
        'ö': 'o', 'Ö': 'O',
        'ç': 'c', 'Ç': 'C'
    }
    
    DEPARTMENT_CODE_MAP = {
        'BLM': 'BILGISAYAR',      # Bilgisayar Mühendisliği
        'BILM': 'BILGISAYAR',
        'YZM': 'YAZILIM',         # Yazılım Mühendisliği
        'MAT': 'MAT',             # Matematik Mühendisliği
        'SEC': 'SEC',             # Seçmeli Dersler
        'EE': 'EE',               # Elektrik-Elektronik
        'ELK': 'EE',
        'ME': 'ME',               # Makine Mühendisliği
        'MAK': 'ME',
        'CE': 'CE',               # İnşaat Mühendisliği
        'INS': 'CE',
    }
    
    DEFAULT_DEPARTMENT_CODE = 'BILGISAYAR'
    
    def __init__(self):
        self.student_repo = StudentRepository()
        self.student_course_repo = StudentCourseRepository()
        self.course_repo = CourseRepository()
        self.department_repo = DepartmentRepository()
        self.lecturer_repo = LecturerRepository()
    
    def import_from_excel(
        self,
        file_path: str,
        course_id: Optional[int] = None,
        course_code: Optional[str] = None,
        department_id: Optional[int] = None,
        semester: Optional[str] = None,
        year: Optional[int] = None
    ) -> ImportResult:
        if not os.path.exists(file_path):
            return ImportResult(
                success=False,
                message=f"Dosya bulunamadı: {file_path}"
            )
        
        target_course_id = course_id
        if not target_course_id and course_code:
            course = self.course_repo.get_by_code(course_code)
            if course:
                target_course_id = course.id
                department_id = department_id or course.department_id
                year = year or course.year
            else:
                logger.info(f"Ders bulunamadı: {course_code}, otomatik oluşturuluyor...")
                created_course = self._create_course_from_code(course_code, file_path)
                if created_course:
                    target_course_id = created_course.id
                    department_id = department_id or created_course.department_id
                    year = year or created_course.year
                    logger.info(f"Ders oluşturuldu: {created_course.code} - {created_course.name}")
                else:
                    return ImportResult(
                        success=False,
                        message=f"Ders bulunamadı ve oluşturulamadı: {course_code}"
                    )
        
        if not target_course_id:
            return ImportResult(
                success=False,
                message="Ders belirtilmedi veya bulunamadı"
            )
        
        try:
            students_data = self._read_excel_file(file_path, department_id, year)
        except Exception as e:
            return ImportResult(
                success=False,
                message=f"Excel dosyası okunamadı: {str(e)}"
            )
        
        if not students_data:
            return ImportResult(
                success=False,
                message="Öğrenci verisi bulunamadı"
            )
        
        result = self._import_students(students_data, target_course_id, semester)
        
        return result
    
    def import_from_excel_directory(
        self,
        directory_path: Optional[str] = None,
        semester: Optional[str] = None,
        department_id: Optional[int] = None,
        auto_detect_course: bool = True
    ) -> Dict[str, ImportResult]:
        if directory_path is None:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            directory_path = os.path.join(base_dir, "..", "exceller")
        
        directory_path = os.path.abspath(directory_path)
        
        if not os.path.exists(directory_path):
            logger.error(f"Klasör bulunamadı: {directory_path}")
            return {'': ImportResult(False, f"Klasör bulunamadı: {directory_path}")}
        
        logger.info(f"Excel dosyaları '{directory_path}' klasöründen taranıyor...")
        
        results = {}
        imported_count = 0
        failed_count = 0
        total_students = 0
        total_errors = []
        
        for filename in os.listdir(directory_path):
            if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
                continue
            
            # Zone.Identifier dosyalarını atla
            if 'Zone.Identifier' in filename or '~$' in filename:
                continue
            
            file_path = os.path.join(directory_path, filename)
            
            course_code = None
            if auto_detect_course:
                course_code = self._extract_course_code_from_filename(filename)
            
            logger.info(f"İçe aktarılıyor: {filename} (Ders Kodu: {course_code or 'Otomatik'})")
            
            result = self.import_from_excel(
                file_path=file_path,
                course_code=course_code,
                department_id=department_id,
                semester=semester
            )
            
            results[filename] = result
            
            if result.success:
                imported_count += 1
                total_students += result.students_imported
            else:
                failed_count += 1
                if result.errors:
                    total_errors.extend(result.errors)
        
        # Öğretim üyesi bilgisini Excel dosyalarından çıkar ve mevcut dersleri güncelle
        logger.info("Mevcut dersler öğretim üyesi bilgisi ile güncelleniyor...")
        self._update_courses_with_lecturer_from_files(directory_path)
        
        summary_msg = f"İçe Aktarma Tamamlandı: {imported_count} başarılı, {failed_count} başarısız, {total_students} öğrenci"
        
        if total_errors:
            summary_msg += f", {len(total_errors)} hata"
            logger.warning(f"İçe aktarma tamamlandı ama hatalar var: {len(total_errors)}")
        else:
            logger.info(summary_msg)
        
        summary_result = ImportResult(
            success=imported_count > 0,
            message=summary_msg,
            students_imported=total_students,
            errors=total_errors if total_errors else None
        )
        
        results['_summary'] = summary_result
        
        return results
    
    def import_all_courses(self, semester: Optional[str] = None) -> ImportResult:
        logger.info("Tüm derslerin içe aktarılması başlatılıyor...")
        
        results = self.import_from_excel_directory(
            directory_path=None,  # Varsayılan exceller klasörü
            semester=semester,
            auto_detect_course=True
        )
        
        # Özet sonucu al
        summary = results.get('_summary')
        
        if summary:
            return summary
        else:
            # Özet yoksa sonuçları manuel hesapla
            imported = sum(1 for r in results.values() if r.success)
            total_students = sum(r.students_imported for r in results.values())
            all_errors = []
            for r in results.values():
                if r.errors:
                    all_errors.extend(r.errors)
            
            return ImportResult(
                success=imported > 0,
                message=f"{imported} ders dosyası içe aktarıldı, {total_students} öğrenci",
                students_imported=total_students,
                errors=all_errors if all_errors else None
            )
    
    def _update_courses_with_lecturer_from_files(self, directory_path: str) -> None:
        """
        Dizindeki Excel dosyalarından öğretim üyesi bilgisini çıkararak,
        mevcut dersleri günceller.
        
        Args:
            directory_path: Excel dosyalarının bulunduğu klasör yolu
        """
        try:
            for filename in os.listdir(directory_path):
                if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
                    continue
                
                # Zone.Identifier dosyalarını atla
                if 'Zone.Identifier' in filename or '~$' in filename:
                    continue
                
                file_path = os.path.join(directory_path, filename)
                
                # Ders kodunu çıkar
                course_code = self._extract_course_code_from_filename(filename)
                if not course_code:
                    continue
                
                # Dersi bul
                course = self.course_repo.get_by_code(course_code)
                if not course:
                    continue
                
                # Excel'den öğretim üyesi bilgisini çıkar
                lecturer_name = self.extract_lecturer_from_file(file_path)
                if not lecturer_name:
                    continue
                
                # Öğretim üyesini bul veya oluştur
                lecturer = self._find_or_create_lecturer(lecturer_name, course.department_id)
                if lecturer:
                    # Dersi güncelle
                    course.lecturer_id = lecturer.id
                    course.lecturer_name = f"{lecturer.first_name} {lecturer.last_name}"
                    self.course_repo.update(course)
                    logger.info(f"Ders güncellendi: {course_code} - Öğretim Üyesi: {course.lecturer_name}")
        
        except Exception as e:
            logger.error(f"Dersler öğretim üyesi bilgisi ile güncellenirken hata: {str(e)}")
    
    def get_available_excel_files(self, directory_path: Optional[str] = None) -> List[Dict[str, str]]:
        if directory_path is None:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            directory_path = os.path.join(base_dir, "..", "exceller")
        
        directory_path = os.path.abspath(directory_path)
        
        if not os.path.exists(directory_path):
            return []
        
        files = []
        for filename in os.listdir(directory_path):
            if not (filename.endswith('.xls') or filename.endswith('.xlsx')):
                continue
            
            if 'Zone.Identifier' in filename or '~$' in filename:
                continue
            
            course_code = self._extract_course_code_from_filename(filename)
            file_path = os.path.join(directory_path, filename)
            
            # Dosya istatistiklerini al
            stat = os.stat(file_path)
            file_size = stat.st_size
            
            files.append({
                'filename': filename,
                'course_code': course_code or 'Belirsiz',
                'path': file_path,
                'size': file_size
            })
        
        # Dosya adına göre sırala
        files.sort(key=lambda x: x['filename'])
        
        return files
    
    def _read_excel_file(
        self,
        file_path: str,
        department_id: Optional[int] = None,
        year: Optional[int] = None
    ) -> List[Dict]:

        students = []
        
        if file_path.endswith('.xlsx'):
            return self._read_xlsx(file_path, department_id, year)
        else:  # .xls
            return self._read_xls(file_path, department_id, year)
    
    def extract_lecturer_from_file(self, file_path: str) -> Optional[str]:
        """
        Excel dosyasından öğretim üyesi bilgisini çıkarır.
        
        Bu dosyalarda öğretim üyesi bilgisi merged cell olarak yer alır.
        Metin içinde desen arayıp ismi çıkarır.
        
        Args:
            file_path: Excel dosya yolu
            
        Returns:
            Öğretim üyesi adı veya None
        """
        try:
            if file_path.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet = workbook.active
            else:  # .xls
                workbook = xlrd.open_workbook(file_path, formatting_info=False)
                sheet = workbook.sheet_by_index(0)
            
            # Öğretim üyesi adı genellikle ilk 5 satırda yer almaktadır
            full_text = ""
            for row_idx in range(min(5, sheet.nrows)):
                # .xls dosyası için
                if file_path.endswith('.xls'):
                    for col_idx in range(sheet.ncols):
                        cell_val = str(sheet.cell_value(row_idx, col_idx))
                        full_text += cell_val + " "
                # .xlsx dosyası için
                else:
                    for cell in list(sheet.rows)[row_idx]:
                        if cell.value:
                            full_text += str(cell.value) + " "
            
            # Metin içinde öğretim üyesi desenini ara
            # Desen: "Dr. Öğr. Üyesi İSIM" veya "Öğr.Gör. İSIM" vb.
            lecturer = self._extract_lecturer_name_from_text(full_text)
            
            if lecturer:
                logger.info(f"Öğretim üyesi bulundu: {lecturer}")
                return lecturer
            
            logger.debug(f"Excel dosyasında öğretim üyesi bilgisi bulunamadı: {file_path}")
            return None
            
        except Exception as e:
            logger.error(f"Öğretim üyesi çıkarma hatası ({file_path}): {str(e)}")
            return None
    
    def _extract_lecturer_name_from_text(self, text: str) -> Optional[str]:
        """
        Metin içinde öğretim üyesi adını çıkarır.
        
        Desenler:
        - "Dr. Öğr. Üyesi ELİF PINAR HACIBEYOĞLU"
        - "Öğr.Gör. ORKUN KARABATAK"
        - "Prof. Dr. AD SOYAD"
        
        Args:
            text: Aranacak metin
            
        Returns:
            Öğretim üyesi adı veya None
        """
        import re
        
        # Türkçe karakterleri normalize et
        def normalize_tr(t: str) -> str:
            return t.replace('İ', 'I').replace('ı', 'i').replace('ş', 's').replace('Ş', 'S').replace('ğ', 'g').replace('Ğ', 'G').replace('ö', 'o').replace('Ö', 'O').replace('ç', 'c').replace('Ç', 'C').replace('ü', 'u').replace('Ü', 'U')
        
        # Desen 1: "Dr. Öğr. Üyesi" veya "Dr. Ogr. Uyesi" + isim (satır sonuna kadar veya "Sınıf Listesi"ne kadar)
        pattern1 = r'(?:Dr\.?\s*Öğr\.?\s*Üyesi|Dr\.?\s*Ogr\.?\s*Uyesi)\s+([A-ZÇĞÖŞÜİ\s]+?)(?:\n|$|Sınıf Listesi)'
        match1 = re.search(pattern1, text, re.IGNORECASE | re.UNICODE)
        if match1:
            name = match1.group(1).strip()
            # Fazla boşlukları temizle
            name = re.sub(r'\s+', ' ', name)
            if len(name) > 3 and len(name) < 100:
                return name
        
        # Desen 2: "Öğr.Gör." + isim
        pattern2 = r'Öğr\.?Gör\.?\s+([A-ZÇĞÖŞÜİ\s]+?)(?:\n|$|Sınıf Listesi)'
        match2 = re.search(pattern2, text, re.IGNORECASE | re.UNICODE)
        if match2:
            name = match2.group(1).strip()
            name = re.sub(r'\s+', ' ', name)
            if len(name) > 3 and len(name) < 100:
                return name
        
        # Desen 3: "Prof. Dr." + isim
        pattern3 = r'Prof\.?\s*Dr\.?\s+([A-ZÇĞÖŞÜİ\s]+?)(?:\n|$|Sınıf Listesi)'
        match3 = re.search(pattern3, text, re.IGNORECASE | re.UNICODE)
        if match3:
            name = match3.group(1).strip()
            name = re.sub(r'\s+', ' ', name)
            if len(name) > 3 and len(name) < 100:
                return name
        
        # Desen 4: "Doç. Dr." + isim
        pattern4 = r'Doç\.?\s*Dr\.?\s+([A-ZÇĞÖŞÜİ\s]+?)(?:\n|$|Sınıf Listesi)'
        match4 = re.search(pattern4, text, re.IGNORECASE | re.UNICODE)
        if match4:
            name = match4.group(1).strip()
            name = re.sub(r'\s+', ' ', name)
            if len(name) > 3 and len(name) < 100:
                return name
        
        # Desen 5: Genel "Dr." + isim (büyük harfle başlayan kelimeler)
        pattern5 = r'Dr\.?\s+([A-ZÇĞÖŞÜİ][a-zA-ZÇĞÖŞÜİçğıöşü]+\s+[A-ZÇĞÖŞÜİ][a-zA-ZÇĞÖŞÜİçğıöşüü]+)'
        match5 = re.search(pattern5, text, re.UNICODE)
        if match5:
            name = match5.group(1).strip()
            name = re.sub(r'\s+', ' ', name)
            if len(name) > 5 and len(name) < 100:
                return name
        
        return None
    
    def _read_xls(
        self,
        file_path: str,
        department_id: Optional[int] = None,
        year: Optional[int] = None
    ) -> List[Dict]:

        students = []
        
        try:
            workbook = xlrd.open_workbook(file_path, formatting_info=False)
            sheet = workbook.sheet_by_index(0)
            
            header_row = self._find_header_row_xls(sheet)
            if header_row is None:
                header_row = 0
            
            col_indices = self._get_column_indices_xls(sheet, header_row)
            
            # Eğer col_indices ile sütun bulunamadıysa, sabit pozisyonları dene
            # Analiz sonucuna göre: Sütun 4 = Öğrenci no, Sütun 5 = Adı Soyadı
            if col_indices.get('student_number') is None:
                col_indices['student_number'] = 4
            if col_indices.get('full_name') is None and col_indices.get('first_name') is None:
                col_indices['full_name'] = 5
            
            for row_idx in range(header_row + 1, sheet.nrows):
                row_data = {}
                
                for key, col_idx in col_indices.items():
                    if col_idx is not None and col_idx < sheet.ncols:
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row_data[key] = self._clean_cell_value(cell_value)
                
                student_number = row_data.get('student_number', '').strip()
                if not student_number or not self._is_valid_student_number(student_number):
                    continue
                
                first_name = row_data.get('first_name', '').strip()
                last_name = row_data.get('last_name', '').strip()
                
                if not first_name and not last_name:
                    full_name = row_data.get('full_name', '').strip()
                    if full_name:
                        name_parts = full_name.split()
                        if len(name_parts) >= 2:
                            first_name = ' '.join(name_parts[:-1])
                            last_name = name_parts[-1]
                        elif len(name_parts) == 1:
                            # Tek isim varsa, adı olarak kullan ve soyadı boş bırak
                            first_name = full_name
                            last_name = "-"  # Boş soyad için yer tutucu
                
                # İsim yoksa satırı atla
                if not first_name:
                    continue
                
                # Soyad yoksa varsayılan değer ata
                if not last_name:
                    last_name = "-"
                
                first_name = first_name[:50]
                last_name = last_name[:50]
                
                student = {
                    'student_number': student_number,
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': row_data.get('email', ''),
                    'department_id': department_id,
                    'year': year or self._parse_year(row_data.get('year', ''))
                }
                
                students.append(student)
                
        except Exception as e:
            raise Exception(f"XLS okuma hatası: {str(e)}")
        
        return students
    
    def _read_xlsx(
        self,
        file_path: str,
        department_id: Optional[int] = None,
        year: Optional[int] = None
    ) -> List[Dict]:
        students = []
        
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active
            
            header_row = self._find_header_row_xlsx(sheet)
            if header_row is None:
                header_row = 1
            
            col_indices = self._get_column_indices_xlsx(sheet, header_row)
            
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                row_data = {}
                
                for key, col_idx in col_indices.items():
                    if col_idx is not None and col_idx < len(row):
                        row_data[key] = self._clean_cell_value(row[col_idx])
                
                student_number = row_data.get('student_number', '').strip()
                if not student_number or not self._is_valid_student_number(student_number):
                    continue
                
                first_name = row_data.get('first_name', '').strip()
                last_name = row_data.get('last_name', '').strip()
                
                if not first_name and not last_name:
                    full_name = row_data.get('full_name', '').strip()
                    if full_name:
                        name_parts = full_name.split()
                        if len(name_parts) >= 2:
                            first_name = ' '.join(name_parts[:-1])
                            last_name = name_parts[-1]
                        else:
                            first_name = full_name
                
                if not first_name or not last_name:
                    continue
                
                # Ad ve soyadı 50 karaktere kırp (database sütun limiti)
                first_name = first_name[:50]
                last_name = last_name[:50]
                
                student = {
                    'student_number': student_number,
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': row_data.get('email', ''),
                    'department_id': department_id,
                    'year': year or self._parse_year(row_data.get('year', ''))
                }
                
                students.append(student)
                
        except Exception as e:
            raise Exception(f"XLSX okuma hatası: {str(e)}")
        
        return students
    
    def _find_header_row_xls(self, sheet) -> Optional[int]:
        # Türkçe karakter normalize etme fonksiyonu
        def normalize_turkish(text: str) -> str:
            # Türkçe karakterleri İngilizce karşılıklarına dönüştür
            tr_map = str.maketrans('ıİğĞşŞöÖçÇüÜ', 'iIgGsSoOcCuU')
            return text.translate(tr_map).lower().strip()
        
        common_headers = [
            'ogrenci no', 'ogrenci no.', 'ogr no', 'ogr no',
            'ogrenci no', 'öğrenci no', 'numara', 'no', 'no.', 'sıra no',
            'student no', 'student number'
        ]
        
        # Başlık satırını bulmak için her satırı kontrol et
        # Başlık satırı "Öğrenci No" ve "Adı Soyadı" gibi temel bilgileri içermeli
        for row_idx in range(min(15, sheet.nrows)):
            row_values = []
            for col_idx in range(min(10, sheet.ncols)):  # İlk 10 sütunu kontrol et
                cell_value = str(sheet.cell_value(row_idx, col_idx)).strip()
                if cell_value:
                    row_values.append(cell_value)
            
            # "Öğrenci no" içeren satırı bul
            row_text = ' '.join(row_values)
            normalized_row = normalize_turkish(row_text)
            
            # En az 2 başlık alanını içermeli
            student_no_found = False
            ad_soyad_found = False
            
            for col_idx in range(sheet.ncols):
                cell_value = str(sheet.cell_value(row_idx, col_idx))
                normalized_value = normalize_turkish(cell_value)
                lower_value = cell_value.lower().strip()
                
                # Öğrenci no kontrolü
                if any(normalize_turkish(h) in normalized_value for h in ['ogrenci no', 'ogr no', 'student no']):
                    student_no_found = True
                # Ad Soyadı kontrolü
                elif any(h in normalized_value for h in ['adi soyadi', 'ad soyad', 'adi ve soyadi']):
                    ad_soyad_found = True
            
            # Her iki alan da bulunduysa bu başlık satırıdır
            if student_no_found and ad_soyad_found:
                return row_idx
        
        return None
    
    def _find_header_row_xlsx(self, sheet) -> Optional[int]:
        common_headers = ['ogrenci no', 'öğrenci no', 'numara', 'no', 'no.', 'sıra no']
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
            for cell_value in row:
                if cell_value:
                    cell_str = str(cell_value).lower().strip()
                    for header in common_headers:
                        if header in cell_str:
                            return row_idx + 1
        
        return None
    
    def _get_column_indices_xls(self, sheet, header_row: int) -> Dict[str, Optional[int]]:
        indices = {
            'student_number': None,
            'first_name': None,
            'last_name': None,
            'full_name': None,
            'email': None,
            'year': None
        }
        
        # Türkçe karakter normalize etme fonksiyonu
        def normalize_turkish(text: str) -> str:
            tr_map = str.maketrans('ıİğĞşŞöÖçÇüÜ', 'iIgGsSoOcCuU')
            return text.translate(tr_map).lower().strip()
        
        if header_row >= sheet.nrows:
            return indices
        
        for col_idx in range(sheet.ncols):
            header = str(sheet.cell_value(header_row, col_idx)).lower().strip()
            normalized_header = normalize_turkish(header)
            
            # Öğrenci numarası kontrolü - öncelik sırasına göre kontrol et
            student_no_patterns = ['ogrenci no', 'ogr no', 'ogr. no', 'numara', 'no', 'student no']
            for pattern in student_no_patterns:
                if pattern in normalized_header and 'ad soyad' not in normalized_header:
                    indices['student_number'] = col_idx
                    break
            
            # Ad Soyadı kontrolü (birleşik isim)
            if any(p in normalized_header for p in ['adi soyadi', 'ad soyad', 'adi ve soyadi']) and indices['full_name'] is None:
                indices['full_name'] = col_idx
            # Ayrı ad kontrolü - sadece 'ad' varsa ve 'ad soyad' yoksa
            elif 'ad' in normalized_header and 'soyad' not in normalized_header and 'ad soyad' not in normalized_header and indices['first_name'] is None:
                indices['first_name'] = col_idx
            # Soyad kontrolü - sadece 'soyad' varsa ve 'ad soyad' yoksa
            elif 'soyad' in normalized_header and 'ad soyad' not in normalized_header and indices['last_name'] is None:
                indices['last_name'] = col_idx
            # E-posta kontrolü
            elif any(p in normalized_header for p in ['e-posta', 'eposta', 'email', 'e-post', 'mail']) and indices['email'] is None:
                indices['email'] = col_idx
            # Sınıf kontrolü
            elif ('sinif' in normalized_header or 'sınıf' in header) and indices['year'] is None:
                indices['year'] = col_idx
        
        return indices
    
    def _get_column_indices_xlsx(self, sheet, header_row: int) -> Dict[str, Optional[int]]:
        indices = {
            'student_number': None,
            'first_name': None,
            'last_name': None,
            'full_name': None,
            'email': None,
            'year': None
        }
        
        header_values = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        
        for col_idx, header in enumerate(header_values):
            if header:
                header_str = str(header).lower().strip()
                
                if any(x in header_str for x in ['öğrenci no', 'ogrenci no', 'numara', 'no.', 'no ']):
                    indices['student_number'] = col_idx
                elif 'ad' in header_str and 'soyad' not in header_str and 'ad soyad' not in header_str:
                    indices['first_name'] = col_idx
                elif 'soyad' in header_str:
                    indices['last_name'] = col_idx
                elif 'ad soyad' in header_str or 'adı soyadı' in header_str:
                    indices['full_name'] = col_idx
                elif any(x in header_str for x in ['e-posta', 'eposta', 'email', 'e-post']):
                    indices['email'] = col_idx
                elif 'sınıf' in header_str:
                    indices['year'] = col_idx
        
        return indices
    
    def _clean_cell_value(self, value) -> str:
        if value is None:
            return ''
        
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        
        return str(value).strip()
    
    def _is_valid_student_number(self, number: str) -> bool:
        if not number:
            return False
        
        number = number.strip()
        if not number.isdigit():
            return False
        
        return len(number) >= 3
    
    def _parse_year(self, year_str: str) -> int:
        if not year_str:
            return 1
        
        year_str = str(year_str).lower().strip()
        
        roman_map = {'i': 1, 'ii': 2, 'iii': 3, 'iv': 4, 'v': 5, 'vi': 6}
        if year_str in roman_map:
            return roman_map[year_str]
        
        if year_str.isdigit():
            year_int = int(year_str)
            return 1 if year_int < 1 or year_int > 6 else year_int
        
        if '1' in year_str or 'bir' in year_str:
            return 1
        elif '2' in year_str or 'iki' in year_str:
            return 2
        elif '3' in year_str or 'üç' in year_str:
            return 3
        elif '4' in year_str or 'dört' in year_str:
            return 4
        
        return 1
    
    def _extract_course_code_from_filename(self, filename: str) -> Optional[str]:
        match = re.search(r'\[([A-Z0-9]+)\]', filename)
        if match:
            return match.group(1)
        
        match = re.search(r'[Cc]lass[Ll]ist[_\s-]+([A-Z]{2,4}\d{3}[A-Za-z]?)', filename)
        if match:
            return match.group(1)
        
        match = re.search(r'([A-Z]{2,4}\d{3}[A-Za-z]?)', filename)
        if match:
            potential_code = match.group(1)
            if len(re.findall(r'[A-Z]', potential_code)) >= 2 and len(re.findall(r'\d', potential_code)) >= 3:
                return potential_code
        
        logger.debug(f"Ders kodu çıkarılamadı: {filename}")
        return None
    
    def _extract_course_name_from_filename(self, filename: str) -> str:
        name = os.path.splitext(filename)[0]
        
        name = re.sub(r'\[[A-Z0-9]+\]', '', name)
        
        name = re.sub(r'\s*\(\d+\)\s*', '', name)
        
        name = name.strip(' _-')
        
        if not name:
            return "Bilinmeyen Ders"
        
        spaced_name = re.sub(r'([a-zışüğöç])([A-ZİŞÜĞÖÇ])', r'\1 \2', name)
        
        # Türkçe karakterleri kontrol et (ve, ile, vs)
        spaced_name = spaced_name.replace('ve', ' ve ').replace('ile', ' ile ')
        
        # Fazla boşlukları temizle
        spaced_name = re.sub(r'\s+', ' ', spaced_name).strip()
        
        return spaced_name if spaced_name else "Bilinmeyen Ders"
    
    def _get_department_from_course_code(self, course_code: str) -> Optional[int]:
        """
        Ders kodundan bölüm ID'sini çıkarır.
        
        Örnek:
        - BLM331 -> 'BILGISAYAR' bölümünün ID'si
        - YZM213 -> 'YAZILIM' bölümünün ID'si
        - MAT213 -> 'MAT' bölümünün ID'si (bulunamazsa varsayılan)
        
        Args:
            course_code: Ders kodu
            
        Returns:
            Bölüm ID'si veya varsayılan bölümün ID'si
        """
        try:
            # Ders kodundan bölüm kodunu çıkar (harfler)
            dept_match = re.match(r'^([A-Z]+)', course_code)
            if not dept_match:
                # Varsayılan bölümü dön
                default_dept = self.department_repo.get_by_code(self.DEFAULT_DEPARTMENT_CODE)
                return default_dept.id if default_dept else None
            
            dept_code = dept_match.group(1)
            
            # DEPARTMENT_CODE_MAP'ten normalize et
            normalized_code = self.DEPARTMENT_CODE_MAP.get(dept_code, dept_code)
            
            # Bölümü bul
            department = self.department_repo.get_by_code(normalized_code)
            if department:
                return department.id
            
            # Bölüm bulunamazsa varsayılan bölümü dön
            logger.warning(f"Bölüm bulunamadı ({normalized_code}), varsayılan bölüm ({self.DEFAULT_DEPARTMENT_CODE}) kullanılıyor")
            default_dept = self.department_repo.get_by_code(self.DEFAULT_DEPARTMENT_CODE)
            return default_dept.id if default_dept else None
            
        except Exception as e:
            logger.error(f"Bölüm ID'si çıkarma hatası ({course_code}): {str(e)}")
            default_dept = self.department_repo.get_by_code(self.DEFAULT_DEPARTMENT_CODE)
            return default_dept.id if default_dept else None
    
    def _get_year_from_course_code(self, course_code: str) -> int:
        """
        Ders kodundan sınıf yılını çıkarır.
        
        Örnek:
        - BLM111 -> 1. sınıf (ilk rakam 1)
        - BLM328 -> 3. sınıf (ilk rakam 3)
        
        Args:
            course_code: Ders kodu
            
        Returns:
            Sınıf yılı (1-4), bulunamazsa 1
        """
        # Ders kodundan ilk rakamı çıkar
        digits_match = re.search(r'(\d)', course_code)
        if digits_match:
            year = int(digits_match.group(1))
            # 1-4 arasında sınırla
            return max(1, min(4, year))
        return 1
    
    def _find_or_create_lecturer(self, lecturer_name: str, department_id: Optional[int] = None) -> Optional[Lecturer]:
        """
        İsimden öğretim üyesini bulur veya yeni oluşturur.
        
        Args:
            lecturer_name: Öğretim üyesi tam adı
            department_id: Bölüm ID'si (opsiyonel)
            
        Returns:
            Lecturer nesnesi veya None
        """
        if not lecturer_name or len(lecturer_name.strip()) < 3:
            return None
        
        lecturer_name = lecturer_name.strip()
        
        # İsim bölme (ad ve soyad)
        name_parts = lecturer_name.split()
        if len(name_parts) < 2:
            # Tek isim varsa, bir unvan olabilir (örn: "Dr. Ahmet")
            first_name = lecturer_name
            last_name = "-"
        else:
            first_name = ' '.join(name_parts[:-1])
            last_name = name_parts[-1]
        
        # İlk olarak var olan öğretim üyelerinde ara
        all_lecturers = self.lecturer_repo.get_all()
        
        # Tam eşleşme ara
        for lecturer in all_lecturers:
            full_name = f"{lecturer.first_name} {lecturer.last_name}".strip()
            if lecturer_name.lower() == full_name.lower():
                logger.info(f"Mevcut öğretim üyesi bulundu: {lecturer_name}")
                return lecturer
        
        # Kısmi eşleşme ara (soyadı kontrol et)
        for lecturer in all_lecturers:
            if lecturer.last_name and lecturer.last_name.lower() == last_name.lower():
                logger.info(f"Öğretim üyesi soyadı eşleşmesi bulundu: {lecturer_name} -> {lecturer}")
                return lecturer
        
        # Yeni öğretim üyesi oluştur
        try:
            new_lecturer = Lecturer(
                first_name=first_name[:50],
                last_name=last_name[:50],
                department_id=department_id,
                title="",  # İsimden unvan çıkarılabilir ama şimdilik boş
                email=None,
                available_days=DEFAULT_AVAILABLE_DAYS.copy()  # Varsayılan müsait günler (liste)
            )
            
            lecturer_id = self.lecturer_repo.create(new_lecturer)
            if lecturer_id:
                new_lecturer.id = lecturer_id
                logger.info(f"Yeni öğretim üyesi oluşturuldu: {lecturer_name} (ID: {lecturer_id})")
                return new_lecturer
            else:
                logger.warning(f"Öğretim üyesi oluşturulamadı: {lecturer_name}")
                return None
                
        except Exception as e:
            logger.error(f"Öğretim üyesi oluşturma hatası ({lecturer_name}): {str(e)}")
            return None
    
    def _create_course_from_code(self, course_code: str, file_path: str) -> Optional[Course]:
        """
        Ders kodundan yeni bir ders oluşturur.
        
        Excel dosyasından öğretim üyesi bilgisini çıkarır ve derse atar.
        
        Args:
            course_code: Ders kodu (örn: BLM331)
            file_path: Excel dosya yolu (ders adını ve öğretim üyesini çıkarmak için)
            
        Returns:
            Oluşturulan Course nesnesi veya None
        """
        try:
            # Dosya adından ders adını çıkar
            filename = os.path.basename(file_path)
            course_name = self._extract_course_name_from_filename(filename)
            
            # Ders kodundan bölüm ID'sini bul
            department_id = self._get_department_from_course_code(course_code)
            
            # Ders kodundan sınıf yılını çıkar
            year = self._get_year_from_course_code(course_code)
            
            # Excel'den öğretim üyesi bilgisini çıkar
            lecturer_name = self.extract_lecturer_from_file(file_path)
            
            # Öğretim üyesini bul veya oluştur
            lecturer_id = None
            lecturer_full_name = None
            if lecturer_name:
                lecturer = self._find_or_create_lecturer(lecturer_name, department_id)
                if lecturer:
                    lecturer_id = lecturer.id
                    lecturer_full_name = f"{lecturer.first_name} {lecturer.last_name}"
            
            # Yeni ders oluştur
            new_course = Course(
                code=course_code,
                name=course_name,
                department_id=department_id,
                lecturer_id=lecturer_id,
                lecturer_name=lecturer_full_name,
                year=year,
                semester=1,  # Varsayılan dönem
                credit=3,    # Varsayılan kredi
                student_count=0,  # Öğrenci sayısı daha sonra güncellenecek
                has_exam=True,
                exam_type='Yazılı',
                exam_duration=90,  # Varsayılan sınav süresi
                course_type='Zorunlu'  # Varsayılan ders tipi
            )
            
            # Veritabanına kaydet
            new_course_id = self.course_repo.create(new_course)
            
            if new_course_id:
                new_course.id = new_course_id
                logger.info(f"Yeni ders oluşturuldu: {course_code} - {course_name} (ID: {new_course_id}, Öğretim Üyesi: {lecturer_full_name or 'Belirsiz'})")
                return new_course
            else:
                logger.error(f"Ders oluşturulamadı: {course_code}")
                return None
                
        except Exception as e:
            logger.error(f"Ders oluşturma hatası ({course_code}): {str(e)}")
            return None
    
    def _import_students(
        self,
        students_data: List[Dict],
        course_id: int,
        semester: Optional[str] = None
    ) -> ImportResult:
        errors = []
        warnings = []
        
        students_to_create = []
        student_courses_to_create = []
        
        for data in students_data:
            try:
                student = Student(
                    student_number=data['student_number'],
                    first_name=data['first_name'],
                    last_name=data['last_name'],
                    email=data.get('email') or None,
                    department_id=data.get('department_id'),
                    year=data.get('year', 1),
                    is_active=True
                )
                
                students_to_create.append(student)
                
            except Exception as e:
                errors.append(f"{data.get('student_number', 'Bilinmeyen')}: {str(e)}")
        
        try:
            created_count = self.student_repo.create_batch(students_to_create)
            
            imported_students = []
            for data in students_data:
                student = self.student_repo.get_by_student_number(data['student_number'])
                if student:
                    imported_students.append(student)
            
            for student in imported_students:
                student_course = StudentCourse(
                    student_id=student.id,
                    course_id=course_id,
                    semester=semester,
                    is_active=True
                )
                student_courses_to_create.append(student_course)
            
            if student_courses_to_create:
                self.student_course_repo.create_batch(student_courses_to_create)
            
            # Dersin öğrenci sayısını güncelle
            student_count = len(imported_students)
            self._update_course_student_count(course_id, student_count)
            
            return ImportResult(
                success=True,
                message=f"{len(imported_students)} öğrenci başarıyla içe aktarıldı",
                students_imported=len(imported_students),
                student_courses_created=len(student_courses_to_create),
                errors=errors if errors else None,
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                message=f"Kayıt sırasında hata: {str(e)}",
                errors=errors
            )
    
    def _update_course_student_count(self, course_id: int, student_count: int) -> bool:
        """
        Dersin öğrenci sayısını günceller.
        
        Args:
            course_id: Ders ID'si
            student_count: Öğrenci sayısı
            
        Returns:
            Güncelleme başarılı mı
        """
        try:
            course = self.course_repo.get_by_id(course_id)
            if course:
                course.student_count = student_count
                self.course_repo.update(course)
                logger.info(f"Ders öğrenci sayısı güncellendi: ID={course_id}, student_count={student_count}")
                return True
            return False
        except Exception as e:
            logger.error(f"Ders öğrenci sayısı güncellenemedi (ID={course_id}): {str(e)}")
            return False
    
    def get_students_by_course(self, course_id: int) -> List[Student]:
        student_courses = self.student_course_repo.get_by_course_id(course_id)
        students = []
        
        for sc in student_courses:
            if sc.student_id:
                student = self.student_repo.get_by_id(sc.student_id)
                if student:
                    students.append(student)
        
        return students
    
    def get_student_course_conflicts(self, course_id1: int, course_id2: int) -> int:
        return self.student_course_repo.check_student_overlap(course_id1, course_id2)
    
    def get_all_conflicts_for_course(self, course_id: int, min_overlap: int = 1) -> List[Dict]:
        return self.student_course_repo.get_conflicting_courses(course_id, min_overlap)
