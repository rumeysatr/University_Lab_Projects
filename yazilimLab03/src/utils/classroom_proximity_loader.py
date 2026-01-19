"""
Derslik Yakınlık Yükleyici
Derslikler arası fiziksel yakınlık bilgilerini Excel/CSV dosyasından okur
ve Graph (Adjacency List) yapısında tutar.

Yeni Excel Formatı (DerslikYakinlik.xlsx):
- Sütun 1: Derslik 1
- Sütun 2: Derslik 2
- Sütun 3: Mesafe (km)
- Sütun 4: Yakınlık Derecesi (1-5, 1=en yakın, 5=en uzak)
"""

import os
import logging
from typing import Dict, List, Set, Optional
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class ClassroomNode:
    name: str
    block: str
    neighbors: Set[str]
    
    def add_neighbor(self, neighbor_name: str) -> None:
        self.neighbors.add(neighbor_name)
    
    def is_neighbor(self, other_name: str) -> bool:
        return other_name in self.neighbors


class ClassroomProximityLoader:
    
    DEFAULT_FILE_PATH = os.path.join("database", "exceller", "DerslikYakinlik.xlsx")
    
    DEFAULT_CSV_PATH = os.path.join("database", "exceller", "DerslikYakinlik.csv")
    
    CAPACITY_FILE_PATH = os.path.join("database", "exceller", "kostu_sinav_kapasiteleri.xlsx")
    
    def __init__(self, file_path: Optional[str] = None):

        self.file_path = file_path or self._resolve_default_path()
        self._classroom_graph: Dict[str, ClassroomNode] = {}
        self._block_classrooms: Dict[str, List[str]] = {}
        self._loaded = False
        self._fallback_warning_logged = False  # Uyarı tekrarını engelle
        
        self._load_data()
    
    def _resolve_default_path(self) -> str:
        # src/utils/ dizininden projenin kök dizinine çık
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        csv_path = os.path.join(base_dir, "database", "exceller", "DerslikYakinlik.csv")
        excel_path = os.path.join(base_dir, "database", "exceller", "DerslikYakinlik.xlsx")
        
        # Path'leri normalize et (Windows/Linux uyumluluğu)
        csv_path = os.path.normpath(csv_path)
        excel_path = os.path.normpath(excel_path)
        
        if os.path.exists(csv_path):
            return csv_path
        return excel_path
    
    def _load_data(self) -> None:
        if self._loaded:
            return
        
        try:
            self._load_from_excel()
        except ImportError:
            try:
                self._load_from_csv()
            except Exception:
                self._load_manual_data()
        except Exception:
            try:
                self._load_from_csv()
            except Exception:
                self._load_manual_data()
    
    def _load_from_excel(self) -> None:
        """Excel dosyasından yeni formatta veri okur."""
        global _global_fallback_warning_logged
        file_path = self._get_actual_file_path()
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Başlık satırını bul
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(h).strip().lower() if h else '' for h in row]
                break
            
            # Sütun indekslerini bul
            col_map = {}
            for idx, header in enumerate(headers):
                if 'derslik 1' in header or 'derslik1' in header or 'classroom1' in header:
                    col_map['classroom1'] = idx
                elif 'derslik 2' in header or 'derslik2' in header or 'classroom2' in header:
                    col_map['classroom2'] = idx
                elif 'mesafe' in header or 'distance' in header:
                    col_map['distance'] = idx
                elif 'yakınlık' in header or 'yakinlik' in header or 'proximity' in header:
                    col_map['proximity'] = idx
            
            # Veri satırlarını oku
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(c is None for c in row):
                    continue
                
                c1 = str(row[col_map.get('classroom1', 0)]).strip() if col_map.get('classroom1') is not None else ''
                c2 = str(row[col_map.get('classroom2', 1)]).strip() if col_map.get('classroom2') is not None else ''
                
                if not c1 or not c2:
                    continue
                
                # Yakınlık derecesine göre komşu olarak ekle (1-3 arası yakındır)
                proximity = 1
                if col_map.get('proximity') is not None:
                    try:
                        proximity = int(row[col_map['proximity']] or 1)
                    except (ValueError, TypeError):
                        proximity = 1
                
                # Sadece yakın derslikleri (1-3) komşu olarak ekle
                if proximity <= 3:
                    self._add_proximity_pair(c1, c2)
            
            wb.close()
            self._loaded = True
            logger.info(f"Yakınlık verileri {file_path} dosyasından yüklendi")
            
        except ImportError:
            # openpyxl yoksa pandas ile dene (eski format)
            self._load_from_pandas()
        except Exception as e:
            if not _global_fallback_warning_logged:
                logger.warning(f"DerslikYakinlik.xlsx dosyası okunamadı, manuel veri kullanılıyor")
                _global_fallback_warning_logged = True
                self._fallback_warning_logged = True
            self._load_manual_data()
    
    def _load_from_pandas(self) -> None:
        """Pandas ile eski formatta veri okur (geriye dönük uyumluluk)."""
        import pandas as pd
        
        file_path = self._get_actual_file_path()
        
        try:
            df = pd.read_excel(file_path, sheet_name="Sayfa1")
            self._parse_dataframe(df)
            self._loaded = True
        except Exception:
            raise FileNotFoundError("Excel dosyası okunamadı")
    
    def _add_proximity_pair(self, classroom1: str, classroom2: str) -> None:
        """İki dersliği karşılıklı komşu olarak ekler."""
        # Classroom1 için
        if classroom1 not in self._classroom_graph:
            node = ClassroomNode(name=classroom1, block=self._extract_block(classroom1), neighbors={classroom2})
            self._classroom_graph[classroom1] = node
        else:
            self._classroom_graph[classroom1].neighbors.add(classroom2)
        
        # Classroom2 için
        if classroom2 not in self._classroom_graph:
            node = ClassroomNode(name=classroom2, block=self._extract_block(classroom2), neighbors={classroom1})
            self._classroom_graph[classroom2] = node
        else:
            self._classroom_graph[classroom2].neighbors.add(classroom1)
        
        # Blok bilgisini güncelle
        block1 = self._extract_block(classroom1)
        block2 = self._extract_block(classroom2)
        
        for block, classroom in [(block1, classroom1), (block2, classroom2)]:
            if block and block not in self._block_classrooms:
                self._block_classrooms[block] = []
            if block and classroom not in self._block_classrooms.get(block, []):
                self._block_classrooms.setdefault(block, []).append(classroom)
    
    def _extract_block(self, classroom_name: str) -> str:
        """Derslik adından blok bilgisini çıkarır."""
        if not classroom_name:
            return ""
        # İlk harf veya harfler bloktur
        for i in range(len(classroom_name)):
            if classroom_name[i].isdigit():
                return classroom_name[:i] if i > 0 else ""
        return classroom_name[0] if classroom_name else ""
    
    def _load_from_csv(self) -> None:
        import csv
        
        try:
            import pandas as pd
            file_path = self._get_actual_file_path()
            
            df = pd.read_excel(file_path, sheet_name="Sayfa1")
            csv_path = self._get_csv_path()
            df.to_csv(csv_path, index=False)
            
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                self._parse_csv_reader(reader)
        except ImportError:
            csv_path = self._get_csv_path()
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                self._parse_csv_reader(reader)
        except Exception:
            raise FileNotFoundError("CSV veya Excel dosyası bulunamadı")
        
        self._loaded = True
    
    def _parse_dataframe(self, df) -> None:
        for _, row in df.iterrows():
            block = str(row.get('BLOK', '')).strip()
            classroom = str(row.get('DERSLİK', '')).strip()
            nearby_str = str(row.get('YAKIN DERSLİK', '')).strip()
            
            if not classroom:
                continue
            
            self._add_classroom_data(block, classroom, nearby_str)
    
    def _parse_csv_reader(self, reader) -> None:
        for row in reader:
            block = row.get('BLOK', '').strip()
            classroom = row.get('DERSLİK', '').strip()
            nearby_str = row.get('YAKIN DERSLİK', '').strip()
            
            if not classroom:
                continue
            
            self._add_classroom_data(block, classroom, nearby_str)
    
    def _add_classroom_data(self, block: str, classroom: str, nearby_str: str) -> None:
        if classroom not in self._classroom_graph:
            node = ClassroomNode(name=classroom, block=block, neighbors=set())
            self._classroom_graph[classroom] = node
        else:
            node = self._classroom_graph[classroom]
            node.block = block  # Blok bilgisini güncelle
        
        if block not in self._block_classrooms:
            self._block_classrooms[block] = []
        if classroom not in self._block_classrooms[block]:
            self._block_classrooms[block].append(classroom)
        
        if nearby_str and nearby_str.lower() != 'nan':
            nearby_list = [n.strip() for n in nearby_str.split(',') if n.strip()]
            for nearby in nearby_list:
                node.add_neighbor(nearby)
                
                if nearby not in self._classroom_graph:
                    nearby_node = ClassroomNode(name=nearby, block='', neighbors=set())
                    self._block_classrooms.setdefault(block, []).append(nearby)
                    self._classroom_graph[nearby] = nearby_node
                self._classroom_graph[nearby].add_neighbor(classroom)
    
    def _load_manual_data(self) -> None:
        manual_data = [
            ("M", "M101", "S101,M201,M301,S201,S202"),
            ("M", "M201", "M301,M101,S201,S202"),
            ("M", "M301", "M201,M101,S201,S202"),
            ("S", "S101", "M101,S201,S202,M201,M301"),
            ("S", "S201", "S101,S202,M201,M301"),
            ("S", "S202", "S101,S201,M201,M301"),
            ("K", "K001", "K002,AMFİA,AMFİB"),
            ("K", "K002", "K001,AMFİA,AMFİB"),
            ("D", "D101", "D102,D103,D104,D201,D202"),
            ("D", "D102", "D101,D103,D104,D201,D202"),
            ("D", "D103", "D101,D102,D104,D201,D202"),
            ("D", "D201", "D202,D101,D103,D104,D102,D301,D302"),
            ("D", "D202", "D201,D101,D103,D104,D102,D301,D302"),
            ("D", "D301", "D201,D202,D302"),
            ("D", "D302", "D201,D202,D301"),
            ("D", "D401", "D301,D302,D403,D402"),
            ("D", "D402", "D301,D302,D403,D401"),
            ("D", "D403", "D301,D302,D402,D401"),
            ("E", "E101", "E102,D201,D202"),
            ("E", "E102", "E101,D201,D202"),
            ("A", "AMFİA", "AMFİB,K001,K002"),
            ("A", "AMFİB", "AMFİA,K001,K002"),
        ]
        
        for block, classroom, nearby_str in manual_data:
            self._add_classroom_data(block, classroom, nearby_str)
        
        self._loaded = True
    
    def _get_actual_file_path(self) -> str:
        """Gerçek dosya yolunu döndürür"""
        # src/utils/ dizininden projenin kök dizinine çık
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        path = os.path.join(base_dir, "database", "exceller", "DerslikYakinlik.xlsx")
        return os.path.normpath(path)
    
    def _get_csv_path(self) -> str:
        # src/utils/ dizininden projenin kök dizinine çık
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        path = os.path.join(base_dir, "database", "exceller", "DerslikYakinlik.csv")
        return os.path.normpath(path)
    
    def get_neighbors(self, classroom_name: str) -> List[str]:
        if classroom_name in self._classroom_graph:
            return list(self._classroom_graph[classroom_name].neighbors)
        return []
    
    def are_neighbors(self, classroom1: str, classroom2: str) -> bool:
        neighbors = self.get_neighbors(classroom1)
        return classroom2 in neighbors
    
    def get_block(self, classroom_name: str) -> str:
        if classroom_name in self._classroom_graph:
            return self._classroom_graph[classroom_name].block
        return ""
    
    def get_classrooms_in_block(self, block: str) -> List[str]:
        return self._block_classrooms.get(block, [])
    
    def get_all_classrooms(self) -> List[str]:
        return list(self._classroom_graph.keys())
    
    def get_available_neighbors_for_combination(
        self, 
        classroom_name: str, 
        available_classrooms: List[str]
    ) -> List[str]:
        all_neighbors = self.get_neighbors(classroom_name)
        available_set = set(available_classrooms)
        
        available_neighbors = [n for n in all_neighbors if n in available_set]
        
        return available_neighbors
    
    def get_closest_classrooms(
        self, 
        classroom_name: str, 
        available_classrooms: List[str],
        limit: int = 5
    ) -> List[str]:
        available_set = set(available_classrooms)
        result = []
        
        direct_neighbors = self.get_neighbors(classroom_name)
        for neighbor in direct_neighbors:
            if neighbor in available_set and neighbor not in result:
                result.append(neighbor)
        
        block = self.get_block(classroom_name)
        if block:
            same_block = self.get_classrooms_in_block(block)
            for other in same_block:
                if other != classroom_name and other in available_set and other not in result:
                    result.append(other)
        
        for other in available_classrooms:
            if other != classroom_name and other not in result:
                result.append(other)
        
        return result[:limit]
    
    def reload(self) -> None:
        self._classroom_graph.clear()
        self._block_classrooms.clear()
        self._loaded = False
        self._load_data()
    
    def is_loaded(self) -> bool:
        return self._loaded
    
    def get_graph_stats(self) -> Dict:
        return {
            'total_classrooms': len(self._classroom_graph),
            'total_blocks': len(self._block_classrooms),
            'blocks': {block: len(classrooms) for block, classrooms in self._block_classrooms.items()},
            'classrooms': list(self._classroom_graph.keys())
        }


_instance: Optional[ClassroomProximityLoader] = None
_global_fallback_warning_logged = False  # Global flag for singleton warning


def get_proximity_loader() -> ClassroomProximityLoader:
    global _instance, _global_fallback_warning_logged
    if _instance is None:
        _instance = ClassroomProximityLoader()
        _global_fallback_warning_logged = _instance._fallback_warning_logged
    return _instance
