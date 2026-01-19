"""
Dışa aktarma controller - Excel ve PDF export işlemleri
View ile Service arasında köprü görevi görür
"""

from typing import List, Dict, Any, Optional, Tuple
from datetime import date, datetime
import os
import logging

from src.services.exam_schedule_service import ExamScheduleService
from src.services.course_service import CourseService
from src.services.classroom_service import ClassroomService
from src.services.student_import_service import StudentImportService
from src.repositories.student_repository import StudentRepository
from src.repositories.course_repository import CourseRepository
from src.utils.excel_generator import ExcelGenerator
from src.utils.pdf_generator import PDFReportGenerator

logger = logging.getLogger(__name__)


def validate_file_path(path: str, extension: str) -> bool:
    if not path:
        return False
    return path.lower().endswith(f'.{extension}')


class ExportController:
    
    def __init__(self, view: Any = None):
        self.view = view
        self.exam_service = ExamScheduleService()
        self.course_service = CourseService()
        self.classroom_service = ClassroomService()
        self.student_repo = StudentRepository()
        self.course_repo = CourseRepository()
        self.excel_generator = ExcelGenerator()
        self.pdf_generator = PDFReportGenerator()
    
    def export_exam_schedule_to_excel(self, start_date: date, end_date: date,
                                      output_path: str = None) -> tuple:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            if output_path and not validate_file_path(output_path, 'xlsx'):
                return False, "Geçersiz dosya yolu veya uzantısı", None
            
            exams = self.exam_service.get_by_date_range(start_date, end_date)
            if not exams:
                return False, "Belirtilen tarih aralığında sınav bulunamadı", None
            
            if not output_path:
                output_path = f"sinav_programi_{start_date}_{end_date}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sınav Programı"
            
            headers = ['Tarih', 'Saat', 'Ders Kodu', 'Ders Adı', 'Öğretim Görevlisi',
                      'Derslik', 'Öğrenci Sayısı', 'Sınav Türü', 'Durum']
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            row = 2
            for exam in exams:
                ws.cell(row=row, column=1, value=exam.exam_date.strftime('%d.%m.%Y'))
                ws.cell(row=row, column=2, value=f"{exam.start_time} - {exam.end_time}")
                ws.cell(row=row, column=3, value=exam.course_code)
                ws.cell(row=row, column=4, value=exam.course_name)
                ws.cell(row=row, column=5, value=exam.lecturer_name)
                ws.cell(row=row, column=6, value=f"{exam.faculty_name or 'Belirsiz'} - {exam.classroom_name}")
                ws.cell(row=row, column=7, value=exam.student_count)
                ws.cell(row=row, column=8, value=self._get_exam_type_label(exam.exam_type))
                ws.cell(row=row, column=9, value=self._get_status_label(exam.status))
                row += 1
            
            column_widths = [15, 15, 15, 30, 25, 20, 12, 15, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            
            wb.save(output_path)
            
            return True, "Excel dosyası başarıyla oluşturuldu", output_path
            
        except ImportError:
            return False, "Excel export için openpyxl kütüphanesi gereklidir", None
        except Exception as e:
            return False, f"Excel oluşturulurken hata: {str(e)}", None
    
    def _get_exam_type_label(self, exam_type: str) -> str:
        labels = {
            'midterm': 'Vize',
            'final': 'Final',
            'makeup': 'Bütünleme',
            'quiz': 'Quiz'
        }
        return labels.get(exam_type, exam_type)
    
    def _get_status_label(self, status: str) -> str:
        labels = {
            'planned': 'Planlandı',
            'confirmed': 'Onaylandı',
            'cancelled': 'İptal Edildi'
        }
        return labels.get(status, status)
    
    def get_export_templates(self) -> Dict[str, str]:
        return {
            'exam_schedule_excel': 'Sınav Programı (Excel)',
        }
    
    def export_to_excel(self, data: list = None, file_path: str = None,
                        filter_type: str = None, filter_value: int = None) -> dict:
        try:
            if data is None:
                if filter_type == 'faculty':
                    data = self.exam_service.get_by_faculty_id(filter_value)
                elif filter_type == 'department':
                    data = self.exam_service.get_by_department_id(filter_value)
                elif filter_type == 'classroom':
                    data = self.exam_service.get_by_classroom_id(filter_value)
                else:
                    data = self.exam_service.get_all()
            
            if not data:
                return {
                    'success': False,
                    'message': 'Dışa aktarılacak veri bulunamadı',
                    'path': None
                }
            
            if not file_path:
                file_path = f"rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            generator = ExcelGenerator()
            success = generator.generate(data, file_path)
            
            return {
                'success': success,
                'message': 'Excel oluşturuldu' if success else 'Excel oluşturulamadı',
                'path': file_path if success else None
            }
            
        except Exception as e:
            return {'success': False, 'message': str(e), 'path': None}
    
    def export_exam_schedule_to_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        output_path: Optional[str] = None
    ) -> Tuple[bool, str, Optional[str]]:
        """
        Sınav programını PDF olarak dışa aktarır.
        
        Args:
            start_date: Başlangıç tarihi
            end_date: Bitiş tarihi
            output_path: Çıktı dosya yolu
            
        Returns:
            (success, message, file_path)
        """
        try:
            if output_path and not validate_file_path(output_path, 'pdf'):
                return False, "Geçersiz dosya yolu veya uzantısı", None
            
            exams = self.exam_service.get_by_date_range(start_date, end_date)
            if not exams:
                return False, "Belirtilen tarih aralığında sınav bulunamadı", None
            
            if not output_path:
                if start_date and end_date:
                    output_path = f"sinav_programi_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
                else:
                    output_path = f"sinav_programi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Excel data formatına çevir
            exam_data = [exam.to_dict() if hasattr(exam, 'to_dict') else exam for exam in exams]
            
            success = self.pdf_generator.generate_exam_schedule_pdf(
                exams=exam_data,
                output_path=output_path,
                start_date=start_date,
                end_date=end_date
            )
            
            if success:
                return True, "PDF dosyası başarıyla oluşturuldu", output_path
            else:
                return False, "PDF oluşturulamadı (reportlab kütüphanesi gerekli)", None
                
        except Exception as e:
            logger.error(f"PDF export hatası: {e}")
            return False, f"PDF oluşturulurken hata: {str(e)}", None
    
    def export_student_schedule_excel(
        self,
        student_number: str,
        output_path: Optional[str] = None
    ) -> Tuple[bool, str, Optional[str]]:
        """
        Belirli bir öğrencinin sınav programını Excel olarak dışa aktarır.
        
        Args:
            student_number: Öğrenci numarası
            output_path: Çıktı dosya yolu
            
        Returns:
            (success, message, file_path)
        """
        try:
            from src.models.student import Student
            
            if not output_path:
                output_path = f"sinav_programi_{student_number}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            student = self.student_repo.get_by_student_number(student_number)
            if not student:
                return False, f"Öğrenci bulunamadı: {student_number}", None
            
            # Öğrencinin derslerini ve sınavlarını al
            from src.repositories.student_course_repository import StudentCourseRepository
            sc_repo = StudentCourseRepository()
            student_courses = sc_repo.get_by_student_id(student.id)
            
            if not student_courses:
                return False, "Öğrencinin kayıtlı dersi bulunamadı", None
            
            # Her ders için sınav programını al
            exam_schedule_data = []
            for sc in student_courses:
                course = self.course_repo.get_by_id(sc.course_id)
                if course:
                    exams = self.exam_service.get_by_course_id(course.id)
                    for exam in exams:
                        exam_info = exam.to_dict() if hasattr(exam, 'to_dict') else exam
                        exam_info['student_name'] = f"{student.first_name} {student.last_name}"
                        exam_info['student_number'] = student.student_number
                        exam_schedule_data.append(exam_info)
            
            if not exam_schedule_data:
                return False, "Öğrenci için planlanmış sınav bulunamadı", None
            
            # Excel oluştur
            success = self.excel_generator.generate(exam_schedule_data, output_path)
            
            if success:
                return True, f"{student_number} numaralı öğrencinin sınav programı dışa aktarıldı", output_path
            else:
                return False, "Excel oluşturulamadı", None
                
        except Exception as e:
            logger.error(f"Öğrenci programı export hatası: {e}")
            return False, f"Dışa aktarılırken hata: {str(e)}", None
    
    def export_course_schedule_excel(
        self,
        course_code: str,
        output_path: Optional[str] = None
    ) -> Tuple[bool, str, Optional[str]]:
        """
        Belirli bir dersin sınav programını Excel olarak dışa aktarır.
        
        Args:
            course_code: Ders kodu
            output_path: Çıktı dosya yolu
            
        Returns:
            (success, message, file_path)
        """
        try:
            if not output_path:
                output_path = f"sinav_programi_{course_code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
            course = self.course_repo.get_by_code(course_code)
            if not course:
                return False, f"Ders bulunamadı: {course_code}", None
            
            exams = self.exam_service.get_by_course_id(course.id)
            if not exams:
                return False, f"{course_code} dersinin sınavı bulunamadı", None
            
            # Excel data formatına çevir
            exam_data = []
            for exam in exams:
                # Öğrencileri de ekle
                from src.repositories.student_course_repository import StudentCourseRepository
                sc_repo = StudentCourseRepository()
                student_courses = sc_repo.get_by_course_id(course.id)
                
                exam_info = exam.to_dict() if hasattr(exam, 'to_dict') else exam
                exam_info['enrolled_students'] = len(student_courses)
                exam_data.append(exam_info)
            
            success = self.excel_generator.generate(exam_data, output_path)
            
            if success:
                return True, f"{course_code} dersinin sınav programı dışa aktarıldı", output_path
            else:
                return False, "Excel oluşturulamadı", None
                
        except Exception as e:
            logger.error(f"Ders programı export hatası: {e}")
            return False, f"Dışa aktarılırken hata: {str(e)}", None
    
    def export_all_courses_report(
        self,
        output_path: Optional[str] = None,
        include_students: bool = True
    ) -> Tuple[bool, str, Optional[str]]:
        """
        Tüm dersler ve öğrencileri içeren kapsamlı rapor Excel olarak dışa aktarır.
        
        Args:
            output_path: Çıktı dosya yolu
            include_students: Öğrencileri de dahil et
            
        Returns:
            (success, message, file_path)
        """
        try:
            if not output_path:
                output_path = f"tum_dersler_raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            
            # Dersler sayfası
            ws_courses = wb.active
            ws_courses.title = "Dersler"
            
            # Ders başlıkları
            course_headers = ['Ders Kodu', 'Ders Adı', 'Öğrenci Sayısı', 'Öğretim Üyesi', 'Bölüm']
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for col, header in enumerate(course_headers, 1):
                cell = ws_courses.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Tüm dersleri al
            courses = self.course_repo.get_all()
            
            row = 2
            for course in courses:
                from src.repositories.student_course_repository import StudentCourseRepository
                sc_repo = StudentCourseRepository()
                student_count = len(sc_repo.get_by_course_id(course.id))
                
                ws_courses.cell(row=row, column=1, value=course.code)
                ws_courses.cell(row=row, column=2, value=course.name)
                ws_courses.cell(row=row, column=3, value=student_count)
                
                lecturer = course.lecturer
                ws_courses.cell(row=row, column=4, value=f"{lecturer.first_name} {lecturer.last_name}" if lecturer else "")
                
                dept = course.department
                ws_courses.cell(row=row, column=5, value=dept.name if dept else "")
                
                for col in range(1, 6):
                    ws_courses.cell(row=row, column=col).border = border
                row += 1
            
            # Sütun genişlikleri
            ws_courses.column_dimensions['A'].width = 15
            ws_courses.column_dimensions['B'].width = 30
            ws_courses.column_dimensions['C'].width = 15
            ws_courses.column_dimensions['D'].width = 20
            ws_courses.column_dimensions['E'].width = 25
            
            if include_students:
                # Öğrenciler sayfası
                ws_students = wb.create_sheet("Öğrenciler")
                
                student_headers = ['Öğrenci No', 'Ad', 'Soyad', 'Bölüm', 'Sınıf', 'Kayıtlı Dersler']
                
                for col, header in enumerate(student_headers, 1):
                    cell = ws_students.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                
                from src.repositories.student_repository import StudentRepository
                student_repo = StudentRepository()
                students = student_repo.get_all()
                
                row = 2
                for student in students:
                    from src.repositories.student_course_repository import StudentCourseRepository
                    sc_repo = StudentCourseRepository()
                    student_courses = sc_repo.get_by_student_id(student.id)
                    
                    course_codes = []
                    for sc in student_courses:
                        c = self.course_repo.get_by_id(sc.course_id)
                        if c:
                            course_codes.append(c.code)
                    
                    ws_students.cell(row=row, column=1, value=student.student_number)
                    ws_students.cell(row=row, column=2, value=student.first_name)
                    ws_students.cell(row=row, column=3, value=student.last_name)
                    
                    dept = student.department
                    ws_students.cell(row=row, column=4, value=dept.name if dept else "")
                    ws_students.cell(row=row, column=5, value=student.year)
                    ws_students.cell(row=row, column=6, value=", ".join(course_codes))
                    
                    for col in range(1, 7):
                        ws_students.cell(row=row, column=col).border = border
                    row += 1
                
                # Sütun genişlikleri
                ws_students.column_dimensions['A'].width = 15
                ws_students.column_dimensions['B'].width = 20
                ws_students.column_dimensions['C'].width = 20
                ws_students.column_dimensions['D'].width = 25
                ws_students.column_dimensions['E'].width = 10
                ws_students.column_dimensions['F'].width = 40
            
            wb.save(output_path)
            
            return True, f"Tüm dersler raporu oluşturuldu ({len(courses)} ders)", output_path
            
        except Exception as e:
            logger.error(f"Genel rapor export hatası: {e}")
            return False, f"Dışa aktarılırken hata: {str(e)}", None
    
    def get_available_exports(self) -> Dict[str, str]:
        """Kullanılabilir export tiplerini döndürür."""
        return {
            'exam_schedule_excel': 'Sınav Programı (Excel)',
            'exam_schedule_pdf': 'Sınav Programı (PDF)',
            'student_schedule_excel': 'Öğrenci Sınav Programı (Excel)',
            'course_schedule_excel': 'Ders Sınav Programı (Excel)',
            'all_courses_report': 'Tüm Dersler Raporu (Excel)',
        }
